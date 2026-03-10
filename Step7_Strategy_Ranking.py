"""
Strategy Correlation Analysis Tool (with MT5 Backtest Ranking)
===============================================================
Scans a folder for SQX trade list .csv files, analyses pairwise correlation,
trade overlap, and produces a formatted Excel report with recommendations.

When --mc-results is provided (recommended), the script first filters strategies
to only those that pass the Monte Carlo 95% confidence Ret/DD threshold,
ensuring only robust strategies enter the correlation analysis.

Strategies are then ranked using metrics parsed from MT5 Strategy Tester HTML reports.

Usage:
    python strategy_correlation_analysis.py [folder_path] [options]

    folder_path               : folder containing *_trades.csv files (default: cwd)
    --mt5-reports FOLDER      : folder containing MT5 .htm reports (default: same as folder_path)
    --mc-results CSV_PATH     : path to BatchMC_Results.csv from QuantAnalyzer
    --mc95-threshold VALUE    : minimum MC95 Ret/DD to pass (default: 2.0)

Examples:
    # With MC filtering (recommended workflow)
    python strategy_correlation_analysis.py C:\\trades --mc-results BatchMC_Results.csv --mc95-threshold 2.5

    # Without MC filtering (analyse all CSVs)
    python strategy_correlation_analysis.py C:\\trades

Requirements:
    pip install pandas openpyxl matplotlib

The script expects SQX trade list exports (.csv) with columns including:
    - Type (Buy/Sell/BuyStop/SellStop/BuyLimit/SellLimit)
    - Open time
    - Close time
    - Profit/Loss

File naming: If filenames end with '_trades' (e.g. SQ_NAS100_H1_1_1_128_trades.csv),
the '_trades' suffix is automatically stripped from the strategy name.

Output:
    Strategy_Correlation_Analysis.xlsx + Dashboard/index.html in the same folder
"""

import sys
import os
import glob
import warnings
import re
import json
import html as html_module
import time
from datetime import datetime
from itertools import combinations

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from html.parser import HTMLParser as _HTMLParser
import zipfile
import xml.etree.ElementTree as ET

warnings.filterwarnings('ignore')

# ============================================================================
# ANSI COLOURS FOR TERMINAL OUTPUT
# ============================================================================
class Colors:
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    GRAY = '\033[90m'
    RESET = '\033[0m'


def init_colors():
    """Enable ANSI colors on Windows."""
    if sys.platform == 'win32':
        os.system('')


def format_duration(seconds: float) -> str:
    """Format duration as hours, minutes, seconds."""
    total_hours = int(seconds // 3600)
    total_minutes = int((seconds % 3600) // 60)
    total_seconds = int(seconds % 60)
    
    if total_hours > 0:
        return f"{total_hours}h {total_minutes}m {total_seconds}s"
    elif total_minutes > 0:
        return f"{total_minutes}m {total_seconds}s"
    else:
        return f"{total_seconds}s"


# ============================================================================
# CONFIGURATION
# ============================================================================
CORR_VERY_HIGH = 0.7
CORR_HIGH = 0.5
CORR_MODERATE = 0.3

OVERLAP_VERY_HIGH = 0.80
OVERLAP_HIGH = 0.60
OVERLAP_MODERATE = 0.40

# ============================================================================
# MT5 RANKING CONFIGURATION
# ============================================================================
MT5_HIGHER_IS_BETTER = {
    "Total Net Profit":   0.15,
    "Ret/DD Ratio":       0.15,
    "Profit Factor":      0.10,
    "Sharpe Ratio":       0.10,
    "Recovery Factor":    0.10,
    "LR Correlation":     0.10,
    "Win/Loss Ratio":     0.05,
    "Total Trades":       0.05,
    "Expected Payoff":    0.05,
    "Win Rate %":         0.05,
}

MT5_LOWER_IS_BETTER = {
    "Balance DD Rel %":   0.05,
    "LR Standard Error":  0.05,
}

MT5_FILE_PATTERNS = ['*.htm', '*.html']

# Default MC95 Ret/DD threshold
DEFAULT_MC95_THRESHOLD = 2.0


# ============================================================================
# MONTE CARLO RESULTS LOADER
# ============================================================================
def load_mc_results(csv_path, threshold=DEFAULT_MC95_THRESHOLD):
    """Load BatchMC_Results.csv and filter strategies by MC95 Ret/DD threshold.

    Returns:
        passing_names: list of normalised strategy names that pass the threshold
        mc95_data: dict of {normalised_name: {mc95_ret_dd, mc95_net_profit, ...}}
        mc_failed: list of dicts for strategies that failed the threshold
    """
    df = pd.read_csv(csv_path, on_bad_lines='skip')
    # Drop completely empty rows
    df = df.dropna(how='all').reset_index(drop=True)
    # Drop rows where Strategy is empty
    df = df.dropna(subset=['Strategy']).reset_index(drop=True)

    # Get MC95 confidence level rows
    mc95 = df[df['ConfidenceLevel'].astype(str) == '95'].copy()

    if mc95.empty:
        print("  WARNING: No 95% confidence level rows found in MC results")
        return [], {}, []

    # Get Original rows for comparison
    orig = df[df['ConfidenceLevel'].astype(str) == 'Original']
    orig_lookup = {}
    for _, row in orig.iterrows():
        orig_lookup[str(row['Strategy']).strip()] = row

    print(f"  Found {len(mc95)} strategies with MC95 data")

    # Apply threshold filter
    mc95['RetDD'] = pd.to_numeric(mc95['RetDD'], errors='coerce').fillna(0)
    passed = mc95[mc95['RetDD'] >= threshold]
    failed = mc95[mc95['RetDD'] < threshold]

    print(f"  MC95 Ret/DD threshold: >= {threshold}")
    print(f"  Passed: {len(passed)}, Failed: {len(failed)}")

    if not failed.empty:
        for _, row in failed.iterrows():
            print(f"    x FILTERED: {row['Strategy']} - MC95 Ret/DD = {row['RetDD']:.2f}")

    def _normalise_mc_name(name):
        """Strip MT5/MT4 suffix and normalise to match trade CSV names."""
        name = str(name).strip()
        for suffix in [' MT5', ' MT4', '_MT5', '_MT4']:
            if name.endswith(suffix):
                name = name[:-len(suffix)]
        name = name.replace('_', ' ')
        return name

    passing_names = []
    mc95_data = {}

    # Store data for ALL MC95 rows (including failed) for reference
    for _, row in mc95.iterrows():
        norm = _normalise_mc_name(row['Strategy'])
        orig_row = orig_lookup.get(str(row['Strategy']).strip())
        entry = {
            'mc95_ret_dd': round(float(row['RetDD']), 2),
            'mc95_net_profit': round(float(row.get('NetProfit', 0)), 2),
            'mc95_max_dd': round(float(row.get('MaxDD', 0)), 2),
            'mc95_max_dd_pct': round(float(row.get('MaxPctDD', 0)), 2),
            'mc95_trades': int(float(row.get('NumTrades', 0))),
            'mc95_r_exp': round(float(row.get('RExp', 0)), 2),
            'mc95_avg_pct_yr': round(float(row.get('AvgPctProfitByYear', 0)), 2),
            'original_name': str(row['Strategy']),
            'orig_net_profit': round(float(orig_row['NetProfit']), 2) if orig_row is not None else 0,
            'orig_ret_dd': round(float(orig_row['RetDD']), 2) if orig_row is not None else 0,
            'orig_trades': int(float(orig_row['NumTrades'])) if orig_row is not None else 0,
        }
        mc95_data[norm] = entry

    for _, row in passed.iterrows():
        norm = _normalise_mc_name(row['Strategy'])
        passing_names.append(norm)

    # Build failed list for dashboard display
    mc_failed = []
    for _, row in failed.iterrows():
        norm = _normalise_mc_name(row['Strategy'])
        d = mc95_data.get(norm, {})
        mc_failed.append({
            'name': str(row['Strategy']),
            'orig_net_profit': d.get('orig_net_profit', 0),
            'mc95_net_profit': d.get('mc95_net_profit', 0),
            'orig_ret_dd': d.get('orig_ret_dd', 0),
            'mc95_ret_dd': d.get('mc95_ret_dd', 0),
            'mc95_max_dd': d.get('mc95_max_dd', 0),
            'mc95_max_dd_pct': d.get('mc95_max_dd_pct', 0),
            'mc95_trades': d.get('mc95_trades', 0),
        })

    return passing_names, mc95_data, mc_failed


def match_csv_to_mc(csv_name, mc_names):
    """Match a trade CSV strategy name to the MC results names.

    Trade CSVs: "SQ NAS100 H1 1.1.102" or "SQ_NAS100_H1_1_1_102"
    MC names:   "SQ NAS100 H1 1.1.102" (already normalised)
    """
    # Normalise the CSV name the same way
    norm = csv_name.replace('_', ' ')
    # Try direct match
    if norm in mc_names:
        return norm
    # Try replacing dots-in-version with different patterns
    # e.g. "SQ NAS100 H1 1 1 102" vs "SQ NAS100 H1 1.1.102"
    import re as _re
    # Collapse spaces and compare
    norm_collapsed = _re.sub(r'\s+', ' ', norm).strip()
    for mc_name in mc_names:
        mc_collapsed = _re.sub(r'\s+', ' ', mc_name).strip()
        if norm_collapsed == mc_collapsed:
            return mc_name
        # Also try matching with dots replaced by spaces
        mc_dotless = mc_collapsed.replace('.', ' ')
        norm_dotless = norm_collapsed.replace('.', ' ')
        if norm_dotless == mc_dotless:
            return mc_name
    return None


# ============================================================================
# DATA LOADING
# ============================================================================
def load_strategies(folder_path):
    csv_files = sorted(glob.glob(os.path.join(folder_path, '*.csv')))
    if not csv_files:
        print(f"ERROR: No .csv files found in '{folder_path}'")
        sys.exit(1)
    # Exclude non-trade files
    csv_files = [f for f in csv_files if 'Correlation_Analysis' not in os.path.basename(f)]
    csv_files = [f for f in csv_files if 'BatchMC_Results' not in os.path.basename(f)]

    strategies = {}
    for filepath in csv_files:
        name = os.path.splitext(os.path.basename(filepath))[0]
        if name.endswith('_trades'):
            name = name[:-7]
        try:
            df = pd.read_csv(filepath)
            required = ['Type', 'Open time', 'Close time', 'Profit/Loss']
            missing = [c for c in required if c not in df.columns]
            if missing:
                print(f"  SKIPPING '{name}': Missing columns {missing}")
                continue
            df['Open time'] = pd.to_datetime(df['Open time'], format='%Y.%m.%d %H:%M:%S')
            df['Close time'] = pd.to_datetime(df['Close time'], format='%Y.%m.%d %H:%M:%S')
            strategies[name] = df
            print(f"  Loaded: {name} ({len(df)} trades)")
        except Exception as e:
            print(f"  SKIPPING '{name}': {e}")

    if len(strategies) < 2:
        print("ERROR: Need at least 2 valid strategy files for correlation analysis.")
        sys.exit(1)
    return strategies


# ============================================================================
# MT5 HTML REPORT PARSER
# ============================================================================
def parse_mt5_report(filepath):
    content = None
    for encoding in ['utf-16-le', 'utf-16', 'utf-8', 'latin-1']:
        try:
            with open(filepath, 'r', encoding=encoding, errors='replace') as f:
                content = f.read()
            if 'Strategy Tester' in content or 'Total Net Profit' in content:
                break
        except Exception:
            continue

    if content is None or 'Total Net Profit' not in content:
        return None

    text = re.sub(r'<[^>]+>', '|', content)
    text = re.sub(r'\|[\s|]*\|', '|', text)
    text = re.sub(r'\|+', '|', text)
    text = re.sub(r'\s+', ' ', text)

    metrics = {}

    def extract_text(label):
        m = re.search(re.escape(label) + r':\|([^|]+)', text)
        return m.group(1).strip() if m else None

    def extract_float(label):
        m = re.search(re.escape(label) + r':\|([\d\s\-\.]+)', text)
        if m:
            try:
                return float(m.group(1).replace(' ', ''))
            except ValueError:
                return None
        return None

    metrics['Expert'] = extract_text('Expert') or os.path.splitext(os.path.basename(filepath))[0]
    metrics['Symbol'] = extract_text('Symbol') or ''
    metrics['Period'] = extract_text('Period') or ''

    metrics['Total Net Profit'] = extract_float('Total Net Profit')
    metrics['Gross Profit'] = extract_float('Gross Profit')
    metrics['Gross Loss'] = extract_float('Gross Loss')
    metrics['Profit Factor'] = extract_float('Profit Factor')
    metrics['Expected Payoff'] = extract_float('Expected Payoff')
    metrics['Recovery Factor'] = extract_float('Recovery Factor')
    metrics['Sharpe Ratio'] = extract_float('Sharpe Ratio')
    metrics['LR Correlation'] = extract_float('LR Correlation')
    metrics['LR Standard Error'] = extract_float('LR Standard Error')

    m = re.search(r'Balance Drawdown Maximal:\|([\d\s\.]+)\(([\d\.]+)%\)', text)
    if m:
        metrics['Balance DD Max $'] = float(m.group(1).replace(' ', ''))
        metrics['Balance DD Max %'] = float(m.group(2))
    # Balance Drawdown Relative (DD as % of balance at peak — matches QA4's DD%)
    m = re.search(r'Balance Drawdown Relative:\|([\d\.]+)%', text)
    if m:
        metrics['Balance DD Rel %'] = float(m.group(1))

    m = re.search(r'Equity Drawdown Maximal:\|([\d\s\.]+)\(([\d\.]+)%\)', text)
    if m:
        metrics['Equity DD Max $'] = float(m.group(1).replace(' ', ''))
        metrics['Equity DD Max %'] = float(m.group(2))

    m = re.search(r'Total Trades:\|(\d+)', text)
    if m: metrics['Total Trades'] = int(m.group(1))

    m = re.search(r'Short Trades \(won %\):\|(\d+) \(([\d\.]+)%\)', text)
    if m:
        metrics['Short Trades'] = int(m.group(1))
        metrics['Short Win %'] = float(m.group(2))

    m = re.search(r'Long Trades \(won %\):\|(\d+) \(([\d\.]+)%\)', text)
    if m:
        metrics['Long Trades'] = int(m.group(1))
        metrics['Long Win %'] = float(m.group(2))

    m = re.search(r'Profit Trades \(% of total\):\|(\d+) \(([\d\.]+)%\)', text)
    if m:
        metrics['Profit Trades'] = int(m.group(1))
        metrics['Win Rate %'] = float(m.group(2))

    m = re.search(r'Loss Trades \(% of total\):\|(\d+)', text)
    if m: metrics['Loss Trades'] = int(m.group(1))

    metrics['Avg Profit Trade'] = extract_float('Average profit trade')
    metrics['Avg Loss Trade'] = extract_float('Average loss trade')
    metrics['Largest Profit Trade'] = extract_float('Largest profit trade')
    metrics['Largest Loss Trade'] = extract_float('Largest loss trade')

    m = re.search(r'Maximum consecutive wins \(\$\):\|(\d+)', text)
    if m: metrics['Max Consecutive Wins'] = int(m.group(1))
    m = re.search(r'Maximum consecutive losses \(\$\):\|(\d+)', text)
    if m: metrics['Max Consecutive Losses'] = int(m.group(1))

    m = re.search(r'Average consecutive wins:\|(\d+)', text)
    if m: metrics['Avg Consecutive Wins'] = int(m.group(1))
    m = re.search(r'Average consecutive losses:\|(\d+)', text)
    if m: metrics['Avg Consecutive Losses'] = int(m.group(1))

    # Ret/DD Ratio: Total Profit / Balance DD Maximal (matches QA4)
    if metrics.get('Total Net Profit') and metrics.get('Balance DD Max $') and metrics['Balance DD Max $'] > 0:
        metrics['Ret/DD Ratio'] = metrics['Total Net Profit'] / metrics['Balance DD Max $']

    # Win/Loss Ratio: Win count / Loss count (matches QA4)
    if metrics.get('Profit Trades') and metrics.get('Loss Trades') and metrics['Loss Trades'] > 0:
        metrics['Win/Loss Ratio'] = metrics['Profit Trades'] / metrics['Loss Trades']

    return metrics


# ============================================================================
# MT5 DEALS TABLE EQUITY EXTRACTOR
# ============================================================================
class _MT5DealsParser(_HTMLParser):
    """Parse MT5 HTML report to extract Deals table rows."""
    def __init__(self):
        super().__init__()
        self.in_row = False
        self.in_cell = False
        self.current_row = []
        self.rows = []

    def handle_starttag(self, tag, attrs):
        if tag == 'tr':
            self.in_row = True
            self.current_row = []
        if tag == 'td' and self.in_row:
            self.in_cell = True

    def handle_endtag(self, tag):
        if tag == 'td':
            self.in_cell = False
        if tag == 'tr' and self.in_row:
            self.in_row = False
            if self.current_row:
                self.rows.append(self.current_row)

    def handle_data(self, data):
        if self.in_cell:
            self.current_row.append(data.strip())


def parse_mt5_deals_equity(filepath):
    """Extract daily equity curve from MT5 HTML report Deals table.

    Finds the Deals table header to locate the Balance and Direction columns,
    then only captures 'out' (close trade) rows for a smooth equity curve.

    Returns list of (datetime, cumulative_profit) or None if parsing fails.
    """
    content = None
    for encoding in ['utf-16-le', 'utf-16', 'utf-8', 'latin-1']:
        try:
            with open(filepath, 'r', encoding=encoding, errors='replace') as f:
                content = f.read()
            if 'Strategy Tester' in content or 'Deals' in content:
                break
        except Exception:
            continue

    if content is None:
        return None

    parser = _MT5DealsParser()
    try:
        parser.feed(content)
    except Exception:
        return None

    if not parser.rows:
        return None

    # Find the Deals table header row containing both "Balance" and "Direction"
    # MT5 Deals: Time, Deal, Symbol, Type, Direction, Volume, Price, Order, Commission, Swap, Profit, Balance, Comment
    deals_start = None
    balance_col = None
    direction_col = None

    for i, row in enumerate(parser.rows):
        if len(row) >= 12:
            bc = dc = None
            for j, cell in enumerate(row):
                s = cell.strip()
                if s == 'Balance':
                    bc = j
                if s == 'Direction':
                    dc = j
            if bc is not None and dc is not None:
                balance_col = bc
                direction_col = dc
                deals_start = i
                break

    if deals_start is None or balance_col is None:
        return None

    # Parse data rows — only "out" direction (close trades) + initial balance row
    deals = []
    initial_balance = None

    for row in parser.rows[deals_start + 1:]:
        # Handle initial balance row (shorter, type="balance")
        if len(row) < 12:
            if len(row) >= 6 and len(row) <= 8:
                type_cell = row[2] if len(row) > 2 else ''
                if type_cell.strip().lower() == 'balance':
                    try:
                        bal = float(row[-1].replace(' ', '').replace('\xa0', ''))
                        initial_balance = bal
                    except (ValueError, TypeError):
                        pass
            continue

        # Only capture close trades (direction = "out")
        direction = row[direction_col].strip().lower() if direction_col < len(row) else ''
        if direction != 'out':
            continue

        # Parse time
        time_str = row[0]
        dt = None
        for fmt in ['%Y.%m.%d %H:%M:%S', '%Y.%m.%d %H:%M', '%Y-%m-%d %H:%M:%S']:
            try:
                dt = pd.to_datetime(time_str, format=fmt)
                break
            except (ValueError, TypeError):
                continue

        if dt is None:
            continue

        # Get Balance value
        try:
            bal_str = row[balance_col].replace(' ', '').replace('\xa0', '')
            if not bal_str:
                continue
            balance = float(bal_str)
        except (ValueError, TypeError, IndexError):
            continue

        if initial_balance is None:
            initial_balance = balance
        deals.append((dt, balance))

    if not deals or initial_balance is None or len(deals) < 5:
        return None

    # Convert to cumulative profit (subtract initial balance)
    equity = [(dt, bal - initial_balance) for dt, bal in deals]

    # Resample to daily — take last value per day
    df = pd.DataFrame(equity, columns=['time', 'profit'])
    df['date'] = df['time'].dt.date
    daily = df.groupby('date')['profit'].last().reset_index()
    daily['date'] = pd.to_datetime(daily['date'])

    return list(zip(daily['date'].tolist(), daily['profit'].tolist()))


def parse_mt5_full_overview(filepath):
    """Parse MT5 HTML report for comprehensive QA4-style overview data.

    Returns dict with:
        - All summary metrics from the report header
        - Monthly P&L grid computed from the Deals table
        - Derived stats: CAGR, Stagnation, Yearly Avg Profit, Win/Loss ratio
    Returns None if parsing fails.
    """
    content = None
    for encoding in ['utf-16-le', 'utf-16', 'utf-8', 'latin-1']:
        try:
            with open(filepath, 'r', encoding=encoding, errors='replace') as f:
                content = f.read()
            if 'Strategy Tester' in content or 'Deals' in content:
                break
        except Exception:
            continue
    if content is None:
        return None

    text = re.sub(r'<[^>]+>', '|', content)
    text = re.sub(r'\|[\s|]*\|', '|', text)
    text = re.sub(r'\|+', '|', text)
    text = re.sub(r'\s+', ' ', text)

    def _ef(label):
        m = re.search(re.escape(label) + r':\|([\d\s\-\.]+)', text)
        if m:
            try: return float(m.group(1).replace(' ', ''))
            except: return None
        return None

    def _et(label):
        m = re.search(re.escape(label) + r':\|([^|]+)', text)
        return m.group(1).strip() if m else None

    o = {}
    o['expert'] = _et('Expert') or os.path.splitext(os.path.basename(filepath))[0]
    o['symbol'] = _et('Symbol') or ''
    o['period'] = _et('Period') or ''
    o['total_profit'] = _ef('Total Net Profit') or 0
    o['gross_profit'] = _ef('Gross Profit') or 0
    o['gross_loss'] = _ef('Gross Loss') or 0
    o['profit_factor'] = _ef('Profit Factor') or 0
    o['expected_payoff'] = _ef('Expected Payoff') or 0
    o['recovery_factor'] = _ef('Recovery Factor') or 0
    o['sharpe'] = _ef('Sharpe Ratio') or 0
    o['lr_corr'] = _ef('LR Correlation') or 0
    o['lr_stderr'] = _ef('LR Standard Error') or 0
    o['total_trades'] = int(_ef('Total Trades') or 0)
    o['avg_win'] = _ef('Average profit trade') or 0
    o['avg_loss'] = _ef('Average loss trade') or 0
    o['largest_win'] = _ef('Largest profit trade') or 0
    o['largest_loss'] = _ef('Largest loss trade') or 0

    # Drawdowns
    m = re.search(r'Balance Drawdown Maximal:\|([\d\s\.]+)\(([\d\.]+)%\)', text)
    if m:
        o['bal_dd_max'] = float(m.group(1).replace(' ', ''))
        o['bal_dd_max_pct'] = float(m.group(2))
    else:
        o['bal_dd_max'] = o['bal_dd_max_pct'] = 0
    # Balance Drawdown Relative is a separate MT5 metric (DD as % of balance at time of DD)
    m = re.search(r'Balance Drawdown Relative:\|([\d\.]+)%', text)
    o['bal_dd_rel_pct'] = float(m.group(1)) if m else o['bal_dd_max_pct']
    m = re.search(r'Equity Drawdown Maximal:\|([\d\s\.]+)\(([\d\.]+)%\)', text)
    if m:
        o['eq_dd_max'] = float(m.group(1).replace(' ', ''))
        o['eq_dd_max_pct'] = float(m.group(2))
    else:
        o['eq_dd_max'] = o['eq_dd_max_pct'] = 0

    # Win/Loss counts
    m = re.search(r'Profit Trades \(% of total\):\|(\d+) \(([\d\.]+)%\)', text)
    o['win_trades'] = int(m.group(1)) if m else 0
    o['win_rate'] = float(m.group(2)) if m else 0
    m = re.search(r'Loss Trades \(% of total\):\|(\d+)', text)
    o['loss_trades'] = int(m.group(1)) if m else 0

    m = re.search(r'Short Trades \(won %\):\|(\d+) \(([\d\.]+)%\)', text)
    o['short_trades'] = int(m.group(1)) if m else 0
    o['short_win_pct'] = float(m.group(2)) if m else 0
    m = re.search(r'Long Trades \(won %\):\|(\d+) \(([\d\.]+)%\)', text)
    o['long_trades'] = int(m.group(1)) if m else 0
    o['long_win_pct'] = float(m.group(2)) if m else 0

    # Consecutive
    m = re.search(r'Maximum consecutive wins \(\$\):\|(\d+)', text)
    o['max_consec_wins'] = int(m.group(1)) if m else 0
    m = re.search(r'Maximum consecutive losses \(\$\):\|(\d+)', text)
    o['max_consec_losses'] = int(m.group(1)) if m else 0
    m = re.search(r'Average consecutive wins:\|(\d+)', text)
    o['avg_consec_wins'] = int(m.group(1)) if m else 0
    m = re.search(r'Average consecutive losses:\|(\d+)', text)
    o['avg_consec_losses'] = int(m.group(1)) if m else 0

    # Derived: Ret/DD uses Balance DD Maximal (matches QA4)
    if o['bal_dd_max'] > 0:
        o['ret_dd'] = round(o['total_profit'] / o['bal_dd_max'], 2)
    else:
        o['ret_dd'] = 0
    # Wins/Losses Ratio = win_count / loss_count (matches QA4)
    if o['loss_trades'] > 0:
        o['wl_ratio'] = round(o['win_trades'] / o['loss_trades'], 2)
    else:
        o['wl_ratio'] = 0
    # Payout Ratio = Avg Win / Avg Loss (separate metric)
    if o['avg_loss'] != 0:
        o['payout_ratio'] = round(abs(o['avg_win'] / o['avg_loss']), 2)
    else:
        o['payout_ratio'] = 0
    # Primary DD display uses Balance DD Maximal $ and Balance DD Relative %
    o['dd_dollar'] = o['bal_dd_max']
    o['dd_pct'] = o['bal_dd_rel_pct']

    # === Parse Deals table for derived stats ===
    parser = _MT5DealsParser()
    try:
        parser.feed(content)
    except Exception:
        pass

    o['monthly_pnl'] = {}
    o['cagr'] = 0
    o['stagnation_days'] = 0
    o['stagnation_pct'] = 0
    o['yearly_avg_profit'] = 0
    o['yearly_avg_return_pct'] = 0
    o['daily_avg_profit'] = 0
    o['monthly_avg_profit'] = 0

    # Find Deals header
    deals_start = balance_col = direction_col = profit_col = None
    for i, row in enumerate(parser.rows):
        if len(row) >= 12:
            bc = dc = pc = None
            for j, cell in enumerate(row):
                s = cell.strip()
                if s == 'Balance': bc = j
                if s == 'Direction': dc = j
                if s == 'Profit': pc = j
            if bc is not None and dc is not None and pc is not None:
                balance_col, direction_col, profit_col = bc, dc, pc
                deals_start = i
                break

    if deals_start is not None:
        trades = []
        initial_balance = None

        for row in parser.rows[deals_start + 1:]:
            if len(row) < 12:
                if 6 <= len(row) <= 8 and len(row) > 2:
                    if row[2].strip().lower() == 'balance':
                        try:
                            initial_balance = float(row[-1].replace(' ', '').replace('\xa0', ''))
                        except (ValueError, TypeError):
                            pass
                continue

            direction = row[direction_col].strip().lower() if direction_col < len(row) else ''
            if direction != 'out':
                continue

            dt = None
            for fmt in ['%Y.%m.%d %H:%M:%S', '%Y.%m.%d %H:%M']:
                try:
                    dt = pd.to_datetime(row[0], format=fmt)
                    break
                except (ValueError, TypeError):
                    continue
            if dt is None:
                continue

            try:
                profit = float(row[profit_col].replace(' ', '').replace('\xa0', ''))
                balance = float(row[balance_col].replace(' ', '').replace('\xa0', ''))
            except (ValueError, TypeError, IndexError):
                continue

            trades.append({'date': dt, 'profit': profit, 'balance': balance})

        if trades and initial_balance:
            df = pd.DataFrame(trades)

            # Monthly P&L grid
            df['year'] = df['date'].dt.year
            df['month'] = df['date'].dt.month
            monthly = df.groupby(['year', 'month'])['profit'].sum()
            years_list = sorted(df['year'].unique())
            monthly_grid = {}
            for yr in years_list:
                row_data = {}
                ytd = 0
                for mo in range(1, 13):
                    val = monthly.get((yr, mo), 0)
                    if isinstance(val, (int, float)):
                        row_data[mo] = round(float(val), 2)
                        ytd += float(val)
                    else:
                        row_data[mo] = 0
                row_data['ytd'] = round(ytd, 2)
                monthly_grid[int(yr)] = row_data
            o['monthly_pnl'] = monthly_grid

            # CAGR
            total_days = (df['date'].max() - df['date'].min()).days
            total_years = total_days / 365.25 if total_days > 0 else 1
            total_profit = df['profit'].sum()
            ending = initial_balance + total_profit
            if initial_balance > 0 and total_years > 0:
                o['cagr'] = round(((ending / initial_balance) ** (1 / total_years) - 1) * 100, 2)

            # Yearly avg profit (total_profit / years — matches QA4)
            if total_years > 0:
                o['yearly_avg_profit'] = round(total_profit / total_years, 2)
            # Yearly avg % return = yearly_avg_profit / initial_balance * 100
            if initial_balance > 0:
                o['yearly_avg_return_pct'] = round(o['yearly_avg_profit'] / initial_balance * 100, 2)

            # Daily avg profit (total_profit / calendar days in period)
            if total_days > 0:
                o['daily_avg_profit'] = round(total_profit / total_days, 2)

            # Monthly avg profit
            monthly_series = df.set_index('date').resample('ME')['profit'].sum()
            if len(monthly_series) > 0:
                o['monthly_avg_profit'] = round(float(monthly_series.mean()), 2)
            else:
                o['monthly_avg_profit'] = 0

            # Compute avg consecutive wins/losses from trade sequence (with decimals)
            wins_runs = []
            losses_runs = []
            current_run = 0
            current_type = None
            for p in df['profit']:
                if p >= 0:
                    if current_type == 'win':
                        current_run += 1
                    else:
                        if current_type == 'loss' and current_run > 0:
                            losses_runs.append(current_run)
                        current_run = 1
                        current_type = 'win'
                else:
                    if current_type == 'loss':
                        current_run += 1
                    else:
                        if current_type == 'win' and current_run > 0:
                            wins_runs.append(current_run)
                        current_run = 1
                        current_type = 'loss'
            # Capture last run
            if current_type == 'win' and current_run > 0:
                wins_runs.append(current_run)
            elif current_type == 'loss' and current_run > 0:
                losses_runs.append(current_run)
            if wins_runs:
                o['avg_consec_wins'] = round(sum(wins_runs) / len(wins_runs), 2)
            if losses_runs:
                o['avg_consec_losses'] = round(sum(losses_runs) / len(losses_runs), 2)

            # Stagnation (longest drawdown period in days)
            equity = df['profit'].cumsum()
            peak = equity.cummax()
            dd = equity - peak
            stag_start = None
            max_stag = 0
            for i in range(len(df)):
                if dd.iloc[i] < 0:
                    if stag_start is None:
                        stag_start = i
                else:
                    if stag_start is not None:
                        days = (df['date'].iloc[i] - df['date'].iloc[stag_start]).days
                        if days > max_stag:
                            max_stag = days
                        stag_start = None
            # Check if still in drawdown at end
            if stag_start is not None:
                days = (df['date'].iloc[-1] - df['date'].iloc[stag_start]).days
                if days > max_stag:
                    max_stag = days

            o['stagnation_days'] = max_stag
            o['stagnation_pct'] = round(max_stag / total_days * 100, 2) if total_days > 0 else 0

            # Annual % / Max DD % ratio (uses Balance DD Relative % to match QA4)
            yearly_avg_return_pct = o['cagr']
            if o['dd_pct'] > 0:
                o['annual_dd_ratio'] = round(yearly_avg_return_pct / o['dd_pct'], 2)
            else:
                o['annual_dd_ratio'] = 0

    return o


def load_mt5_overviews(mt5_folder, strategy_names):
    """Load full MT5 overviews for all matched strategies.

    Returns dict: {strategy_name: overview_dict}
    """
    mt5_report_files = find_mt5_reports(mt5_folder)
    overviews = {}
    for name in strategy_names:
        report_path = match_strategy_to_report(name, mt5_report_files)
        if report_path:
            ov = parse_mt5_full_overview(report_path)
            if ov:
                overviews[name] = ov
    return overviews


# ============================================================================
# STRATEGY PSEUDO CODE PARSER
# ============================================================================

_INDICATOR_NAMES = {
    'MTKeltnerChannel': 'Keltner Channel', 'KeltnerChannel': 'Keltner Channel',
    'HullMovingAverage': 'Hull MA', 'HMA': 'Hull MA',
    'SMMA': 'Smoothed MA (SMMA)',
    'SMA': 'SMA', 'EMA': 'EMA', 'LWMA': 'Linear Weighted MA', 'WMA': 'Weighted MA',
    'KAMA': 'Kaufman Adaptive MA', 'KaufmanEfficiencyRatio': 'Kaufman Efficiency Ratio',
    'KER': 'Kaufman Efficiency Ratio',
    'SuperTrend': 'SuperTrend', 'ADX': 'ADX',
    'LinReg': 'Linear Regression', 'BB Width Ratio': 'BB Width Ratio',
    'BB Range': 'BB Range', 'BollingerBands': 'Bollinger Bands',
    'UlcerIndex': 'Ulcer Index',
    'Williams % R': "Williams %R", 'WilliamsPR': "Williams %R",
    'Pivots': 'Pivot Points', 'RSI': 'RSI', 'CCI': 'CCI',
    'Stochastic': 'Stochastic', 'MACD': 'MACD', 'Ichimoku': 'Ichimoku',
    'DonchianChannel': 'Donchian Channel', 'ParabolicSAR': 'Parabolic SAR',
    'DemarkerIndicator': 'DeMarker', 'Demarker': 'DeMarker', 'DeMarker': 'DeMarker',
    'AO': 'Awesome Oscillator', 'AwesomeOscillator': 'Awesome Oscillator',
    'AC': 'Accelerator Oscillator',
    'MFI': 'Money Flow Index', 'OBV': 'On Balance Volume',
    'ForceIndex': 'Force Index', 'RVI': 'RVI', 'Momentum': 'Momentum',
    'ROC': 'Rate of Change', 'TRIX': 'TRIX',
    'StandardDeviation': 'Std Deviation', 'StdDev': 'Std Deviation',
    'Envelopes': 'Envelopes', 'FractalDimension': 'Fractal Dimension',
    'ChaikinOscillator': 'Chaikin Oscillator', 'ChaikinVolatility': 'Chaikin Volatility',
    'AroonOscillator': 'Aroon', 'ZigZag': 'ZigZag',
    'VWAP': 'VWAP', 'SRPercentRank': 'SR Percent Rank',
    'BullsPower': 'Bulls Power', 'BearsPower': 'Bears Power',
    'HeikenAshiHigh': 'Heiken Ashi', 'HeikenAshiLow': 'Heiken Ashi',
    'HeikenAshiOpen': 'Heiken Ashi', 'HeikenAshiClose': 'Heiken Ashi',
    'MTATR': 'MT ATR', 'Highest': 'Highest', 'Lowest': 'Lowest',
    'GannHiLo': 'Gann HiLo', 'QQE': 'QQE',
    'FractalUp': 'Fractals', 'FractalDown': 'Fractals',
    'Vortex': 'Vortex', 'Reflex': 'Reflex', 'Cycle': 'Cycle',
    'Fibo': 'Fibonacci', 'BiggestRange': 'Biggest Range',
    'SmallestRange': 'Smallest Range',
    'TrueRange': 'True Range', 'BarRange': 'Bar Range',
    'Plus': '+DI', 'Minus': '-DI',
    'IsUptrend': 'Is Uptrend', 'IsDowntrend': 'Is Downtrend',
    'OSMA': 'OsMA (MACD Histogram)',
}

_INDICATOR_PATTERNS = [
    r'(MTKeltnerChannel|KeltnerChannel)\b',
    r'(HullMovingAverage)\b', r'(?<!\w)(HMA)\b', r'(SMMA)\b', r'(SuperTrend)\b',
    r'(KaufmanEfficiencyRatio)\b', r'(?<!\w)(KER)\b', r'(?<!\w)(ADX)\b',
    r'(LinReg)\b', r'(BB Width Ratio|BB Range|BollingerBands)\b',
    r'(UlcerIndex)\b', r'(Williams\s*%\s*R)\b', r'(Pivots)\b',
    r'(?<!\w)(KAMA)\b', r'(LWMA)\b', r'(?<!\w)(WMA)\b',
    r'(?<!\w)(EMA)\b', r'(?<!\w)(SMA)\b',
    r'(?<!\w)(RSI)\b', r'(CCI)\b', r'(Stochastic)\b', r'(MACD)\b',
    r'(Ichimoku)\b', r'(DonchianChannel)\b', r'(ParabolicSAR)\b',
    r'(DemarkerIndicator|Demarker|DeMarker)\b',
    r'(?<!\w)(AO)\b', r'(AwesomeOscillator)\b', r'(?<!\w)(AC)\b',
    r'(MFI)\b', r'(OBV)\b', r'(ForceIndex)\b', r'(?<!\w)(RVI)\b',
    r'(Momentum)\b', r'(ROC)\b', r'(TRIX)\b',
    r'(StandardDeviation|StdDev)\b', r'(Envelopes)\b',
    r'(FractalDimension)\b', r'(ChaikinOscillator|ChaikinVolatility)\b',
    r'(AroonOscillator)\b', r'(ZigZag)\b',
    r'(?<!\w)(VWAP)\b', r'(SRPercentRank)\b',
    r'(BullsPower)\b', r'(BearsPower)\b',
    r'(HeikenAshiHigh|HeikenAshiLow|HeikenAshiOpen|HeikenAshiClose)\b',
    r'(?<!\w)(MTATR)\b',
    r'(?<!\w)(Highest|Lowest)\(Main chart',
    r'(GannHiLo)\b', r'(?<!\w)(QQE)\b',
    r'(FractalUp|FractalDown)\b',
    r'(?<!\w)(Vortex)\b', r'(?<!\w)(Reflex)\b', r'(?<!\w)(Cycle)\b',
    r'(?<!\w)(Fibo)\b',
    r'(BiggestRange)\b', r'(SmallestRange)\b',
    r'(?<!\w)(TrueRange)\b', r'(?<!\w)(BarRange)\b',
    r'(?<!\w)(IsUptrend|IsDowntrend)\b',
    r'(?<!\w)(OSMA)\b',
]

_REFERENCE_PATTERNS = [
    r'(HighestInRange|LowestInRange)\b',
    r'(HighMonthly|LowMonthly|HighWeekly|LowWeekly|HighDaily|LowDaily)\b',
    r'(CloseDaily|OpenDaily|OpenMonthly|CloseMonthly|OpenWeekly|CloseWeekly)\b',
    r'(SessionClose|SessionOpen)\b',
    r'(SessionHigh|SessionLow)\b',
]

_REFERENCE_NAMES = {
    'HighestInRange': 'Session High', 'LowestInRange': 'Session Low',
    'HighMonthly': 'Monthly High', 'LowMonthly': 'Monthly Low',
    'HighWeekly': 'Weekly High', 'LowWeekly': 'Weekly Low',
    'HighDaily': 'Daily High', 'LowDaily': 'Daily Low',
    'CloseDaily': 'Daily Close', 'OpenDaily': 'Daily Open',
    'OpenMonthly': 'Monthly Open', 'CloseMonthly': 'Monthly Close',
    'OpenWeekly': 'Weekly Open', 'CloseWeekly': 'Weekly Close',
    'SessionClose': 'Session Close', 'SessionOpen': 'Session Open',
    'SessionHigh': 'Session High', 'SessionLow': 'Session Low',
}


def parse_strategy_pseudo_code(filepath):
    """Parse a StrategyQuant X pseudo source code .txt file."""
    try:
        content = open(filepath, 'r', errors='replace').read()
    except Exception:
        return None

    o = {}
    m = re.search(r'Pseudo Source Code of (.+)', content)
    o['code_name'] = m.group(1).strip() if m else os.path.basename(filepath)
    m = re.search(r'Backtested on (.+?)/', content)
    o['symbol'] = m.group(1).strip() if m else ''
    m = re.search(r'Backtested on .+?/\s*(\w+)', content)
    o['timeframe'] = m.group(1).strip() if m else ''

    has_long = bool(re.search(r'Open Long order', content))
    has_short = bool(re.search(r'Open Short order', content))
    if has_long and has_short:
        o['direction'] = 'Long & Short'
    elif has_long:
        o['direction'] = 'Long Only'
    elif has_short:
        o['direction'] = 'Short Only'
    else:
        o['direction'] = 'Unknown'

    if ' Stop;' in content:
        o['entry_type'] = 'Stop'
    elif ' Limit;' in content:
        o['entry_type'] = 'Limit'
    else:
        o['entry_type'] = 'Market'

    o['trailing_stop'] = bool(re.search(r'Trailing Stop\s*=', content))
    o['ts_activation'] = bool(re.search(r'TS Activation', content))
    o['move_sl_be'] = bool(re.search(r'Move SL to BE', content))
    o['profit_target'] = bool(re.search(r'Profit target\s*=', content))

    m = re.search(r'Exit After (\w+) bars', content)
    if m:
        val = m.group(1)
        if not val.isdigit():
            pm = re.search(rf'int {val}\s*=\s*(\d+)', content)
            val = pm.group(1) if pm else val
        o['exit_after_bars'] = val
    else:
        o['exit_after_bars'] = None

    long_exit = re.search(r'LongExitSignal\s*=\s*(.+?);', content, re.DOTALL)
    short_exit = re.search(r'ShortExitSignal\s*=\s*(.+?);', content, re.DOTALL)
    o['has_exit_signals'] = (
        (long_exit is not None and long_exit.group(1).strip().lower() != 'false')
        or (short_exit is not None and short_exit.group(1).strip().lower() != 'false')
    )

    m = re.search(r'LimitSignalsTimeRange\s*=\s*true\s*\(([^)]+)\)', content)
    o['time_filter'] = m.group(1).strip() if m else None
    m = re.search(r'Order valid for (\d+) bars', content)
    o['order_valid_bars'] = int(m.group(1)) if m else None

    # Indicators
    indicators = []
    seen = set()
    for pattern in _INDICATOR_PATTERNS:
        for m in re.finditer(pattern, content):
            raw = m.group(1).strip()
            friendly = _INDICATOR_NAMES.get(raw, raw)
            if friendly not in seen:
                seen.add(friendly)
                indicators.append(friendly)
    o['indicators'] = indicators

    # Reference levels
    refs = []
    ref_seen = set()
    for pattern in _REFERENCE_PATTERNS:
        for m in re.finditer(pattern, content):
            raw = m.group(1).strip()
            friendly = _REFERENCE_NAMES.get(raw, raw)
            if friendly not in ref_seen:
                ref_seen.add(friendly)
                refs.append(friendly)
    o['entry_refs'] = refs

    # Style classification
    all_text = content
    breakout_clues = [
        r'KeltnerChannel.*(?:Upper|Lower)', r'HighestInRange|LowestInRange',
        r'HighMonthly|LowMonthly|HighWeekly|LowWeekly|HighDaily|LowDaily',
        r'Pivots', r'SuperTrend', r'DonchianChannel',
        r'SessionHigh|SessionLow', r'GannHiLo',
        r'FractalUp|FractalDown',
    ]
    breakout_score = sum(1 for p in breakout_clues if re.search(p, all_text))
    mr_clues = [
        r'Williams\s*%\s*R', r'(?<!\w)(RSI|Stochastic|CCI)\b',
        r'BB Width|BollingerBands', r'overbought|oversold',
        r'SmallestRange',
    ]
    mr_score = sum(1 for p in mr_clues if re.search(p, all_text, re.IGNORECASE))
    tf_clues = [
        r'(?<!\w)(SMMA|SMA|EMA|LWMA|WMA|HullMovingAverage|HMA|KAMA)\b', r'(?<!\w)ADX\b',
        r'LinReg', r'is above|is below|crossed above|crossed below',
        r'Ichimoku', r'GannHiLo', r'Vortex',
        r'IsUptrend|IsDowntrend',
    ]
    tf_score = sum(1 for p in tf_clues if re.search(p, all_text))
    vol_clues = [r'UlcerIndex', r'KaufmanEfficiencyRatio', r'FractalDimension', r'KER\b']
    vol_score = sum(1 for p in vol_clues if re.search(p, all_text))
    mom_clues = [r'is rising for \d+ bars', r'is falling for \d+ bars',
                 r'BullsPower|BearsPower', r'changes direction',
                 r'QQE', r'Reflex', r'Cycle']
    mom_score = sum(1 for p in mom_clues if re.search(p, all_text))

    if o['entry_type'] == 'Stop':
        if re.search(r'Open Long.*(?:Upper|Highest|High)', content):
            breakout_score += 2
    elif o['entry_type'] == 'Limit':
        mr_score += 2

    styles = []
    if breakout_score >= 2 and breakout_score >= mr_score:
        styles.append('Breakout')
    if mr_score >= 2 and mr_score > breakout_score:
        styles.append('Mean Reversion')
    if tf_score >= 2:
        styles.append('Trend Following')
    if vol_score >= 1 and not styles:
        styles.append('Volatility-Based')
    if mom_score >= 1 and 'Breakout' not in styles and 'Mean Reversion' not in styles:
        styles.append('Momentum')
    if not styles:
        if o['entry_type'] == 'Stop':
            styles.append('Breakout')
        elif o['entry_type'] == 'Limit':
            styles.append('Mean Reversion')
        else:
            styles.append('Systematic')
    o['style'] = ' / '.join(styles)

    exits = []
    if o['trailing_stop']:
        ts = 'Trailing Stop'
        if o['ts_activation']:
            ts += ' (with activation level)'
        exits.append(ts)
    if o['move_sl_be']:
        exits.append('Move SL to Breakeven')
    if o['profit_target']:
        exits.append('Profit Target')
    if o['exit_after_bars']:
        exits.append(f'Exit After {o["exit_after_bars"]} Bars')
    if o['has_exit_signals']:
        exits.append('Signal-Based Exit')
    exits.append('Stop Loss')
    o['exit_summary'] = ', '.join(exits)

    return o


def load_strategy_pseudo_codes(folder_path, strategy_names):
    """Load strategy pseudo code .txt files for all strategies."""
    txt_files = {}
    for root, dirs, files in os.walk(folder_path):
        for f in files:
            if f.lower().endswith('.txt'):
                txt_files[f] = os.path.join(root, f)

    results = {}
    for name in strategy_names:
        norm = name.replace(' ', '_').replace('(', '_').replace(')', '_').replace('.', '_')
        matched = None
        for fname, fpath in txt_files.items():
            fnorm = fname.replace(' ', '_').replace('(', '_').replace(')', '_').replace('.', '_')
            fnorm_base = fnorm.rsplit('_txt', 1)[0] if fnorm.endswith('_txt') else fnorm
            if norm in fnorm_base or fnorm_base in norm:
                matched = fpath
                break
        if not matched:
            m = re.search(r'(\d+[._]\d+[._]\d+(?:[._]\d+)?)', name)
            if m:
                ver = m.group(1).replace('.', '_')
                for fname, fpath in txt_files.items():
                    if ver in fname.replace('.', '_'):
                        matched = fpath
                        break
        if matched:
            parsed = parse_strategy_pseudo_code(matched)
            if parsed:
                # Also store the raw file content for export
                try:
                    with open(matched, 'r', errors='replace') as rf:
                        parsed['raw_content'] = rf.read()
                except Exception:
                    parsed['raw_content'] = ''
                results[name] = parsed
    if results:
        print(f"  Parsed {len(results)} strategy pseudo code files")
    return results


def load_sqx_metadata(folder_path, strategy_names):
    """Load metadata from .sqx files (which are ZIP archives containing settings.xml).

    Currently extracts:
      - Complexity (integer): total number of parameter combinations for a strategy

    Returns dict: {strategy_name: {complexity: int, ...}}
    """
    # Collect all .sqx files in the folder tree
    sqx_files = {}
    for root, dirs, files in os.walk(folder_path):
        for f in files:
            if f.lower().endswith('.sqx'):
                sqx_files[f] = os.path.join(root, f)

    if not sqx_files:
        return {}

    results = {}
    for sqx_name, sqx_path in sqx_files.items():
        try:
            with zipfile.ZipFile(sqx_path, 'r') as zf:
                if 'settings.xml' not in zf.namelist():
                    continue
                xml_content = zf.read('settings.xml').decode('utf-8', errors='replace')

                # Extract strategy name from StrategyName element
                strat_name = None
                name_match = re.search(
                    r'<StrategyName\s+type="String">([^<]+)</StrategyName>',
                    xml_content
                )
                if name_match:
                    strat_name = name_match.group(1).strip()

                # Extract Complexity
                complexity = None
                comp_match = re.search(
                    r'<Complexity\s+type="Integer">(\d+)</Complexity>',
                    xml_content
                )
                if comp_match:
                    complexity = int(comp_match.group(1))

                if strat_name and complexity is not None:
                    results[strat_name] = {
                        'complexity': complexity,
                    }
        except (zipfile.BadZipFile, KeyError, Exception) as e:
            print(f"  WARNING: Could not read {sqx_name}: {e}")
            continue

    # Match to strategy_names (handle variant suffixes like "(1)")
    matched = {}
    for name in strategy_names:
        # Direct match
        if name in results:
            matched[name] = results[name]
            continue
        # Try matching base name (strip trailing " (N)" suffix)
        base = re.sub(r'\s*\(\d+\)\s*$', '', name)
        if base in results:
            matched[name] = results[base]
            continue
        # Fuzzy: normalize and check version numbers
        norm = name.replace(' ', '_').replace('(', '_').replace(')', '_').replace('.', '_')
        for sqx_strat_name, meta in results.items():
            sqx_norm = sqx_strat_name.replace(' ', '_').replace('(', '_').replace(')', '_').replace('.', '_')
            if norm in sqx_norm or sqx_norm in norm:
                matched[name] = meta
                break

    if matched:
        print(f"  Loaded SQX metadata for {len(matched)} strategies")
    return matched


def generate_equity_charts(folder_path, strategies, names, mt5_folder=None):
    """Generate SQX vs MT5 equity comparison chart PNGs for each strategy.

    For each strategy:
    - SQX equity: cumulative P&L from the trade CSV (already loaded in strategies dict)
    - MT5 equity: parsed from MT5 HTML report Deals table Balance column

    Charts are saved as PNGs in the Dashboard/ subfolder.
    Returns dict: {strategy_name: relative_path_to_png}
    """
    dashboard_dir = os.path.join(folder_path, 'Dashboard')
    os.makedirs(dashboard_dir, exist_ok=True)

    mt5_reports_folder = mt5_folder or folder_path
    mt5_report_files = find_mt5_reports(mt5_reports_folder)

    chart_map = {}

    for name in names:
        df = strategies[name]

        # Build SQX equity from CSV trade data
        df_sorted = df.sort_values('Close time').reset_index(drop=True)
        sqx_dates = df_sorted['Close time'].values
        sqx_equity = df_sorted['Profit/Loss'].cumsum().values

        # Resample SQX to daily
        sqx_df = pd.DataFrame({'time': df_sorted['Close time'], 'profit': df_sorted['Profit/Loss'].cumsum()})
        sqx_df['date'] = sqx_df['time'].dt.date
        sqx_daily = sqx_df.groupby('date')['profit'].last().reset_index()
        sqx_daily['date'] = pd.to_datetime(sqx_daily['date'])

        # Try to find and parse MT5 equity
        mt5_equity = None
        report_path = match_strategy_to_report(name, mt5_report_files)
        if report_path:
            mt5_equity = parse_mt5_deals_equity(report_path)

        # Generate chart
        fig, ax = plt.subplots(figsize=(8, 3.5), dpi=130)
        fig.patch.set_facecolor('#12151e')
        ax.set_facecolor('#12151e')

        # Plot SQX equity
        ax.plot(sqx_daily['date'], sqx_daily['profit'],
                color='#6c9eff', linewidth=1.3, label='StrategyQuant X', alpha=0.9)

        # Plot MT5 equity if available
        if mt5_equity and len(mt5_equity) > 1:
            mt5_dates = [d for d, _ in mt5_equity]
            mt5_vals = [v for _, v in mt5_equity]
            ax.plot(mt5_dates, mt5_vals,
                    color='#fbbf24', linewidth=1.3, label='MetaTrader 5', alpha=0.9)

        # Styling
        ax.set_title(name, color='#e8eaf0', fontsize=10, fontweight='bold', pad=10,
                      fontfamily='sans-serif')
        ax.legend(loc='upper left', fontsize=7.5, framealpha=0.3,
                  facecolor='#1e2333', edgecolor='#2a2f42',
                  labelcolor='#8b90a5')

        ax.tick_params(colors='#5c6178', labelsize=7)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_color('#2a2f42')
        ax.spines['left'].set_color('#2a2f42')
        ax.grid(True, alpha=0.15, color='#353b52', linewidth=0.5)
        ax.set_ylabel('Cumulative P&L ($)', color='#8b90a5', fontsize=7.5)

        # Format x-axis dates
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
        ax.xaxis.set_major_locator(mdates.YearLocator())
        fig.autofmt_xdate(rotation=0)

        # Zero line
        ax.axhline(y=0, color='#5c6178', linewidth=0.5, linestyle='--', alpha=0.5)

        plt.tight_layout()

        # Save
        safe_name = re.sub(r'[^\w\-.]', '_', name)
        png_path = os.path.join(dashboard_dir, f'{safe_name}_equity.png')
        fig.savefig(png_path, facecolor='#12151e', edgecolor='none',
                    bbox_inches='tight', pad_inches=0.1)
        plt.close(fig)

        chart_map[name] = f'{safe_name}_equity.png'
        has_mt5 = 'yes' if mt5_equity else 'no'
        print(f"    Chart: {name} (MT5: {has_mt5})")

    return chart_map


def find_mt5_reports(mt5_folder):
    files = []
    for pattern in MT5_FILE_PATTERNS:
        files.extend(glob.glob(os.path.join(mt5_folder, pattern)))
    return sorted(files)


def match_strategy_to_report(strategy_name, mt5_reports):
    def normalise(s):
        # If it looks like a filepath, extract basename and strip extension
        if os.sep in s or '/' in s or s.endswith('.htm') or s.endswith('.html'):
            s = os.path.splitext(os.path.basename(s))[0]
        s = s.lower()
        # Replace underscores and dots with spaces first
        s = re.sub(r'[_.]', ' ', s)
        s = re.sub(r'\s+', ' ', s).strip()
        # Strip common suffixes AFTER normalising separators
        for suffix in [' mt5', ' mt4', ' backtest', ' report']:
            if s.endswith(suffix):
                s = s[:-len(suffix)].strip()
        return s

    norm_name = normalise(strategy_name)

    # Exact match first
    for report_path in mt5_reports:
        norm_report = normalise(report_path)
        if norm_name == norm_report:
            return report_path

    # Fuzzy fallback: check containment, but only if the match is significant
    # (avoid "3 1 182" matching "3 1 186" via substring)
    best_match = None
    best_len = 0
    for report_path in mt5_reports:
        norm_report = normalise(report_path)
        if norm_name in norm_report:
            if len(norm_name) > best_len:
                best_match = report_path
                best_len = len(norm_name)
        elif norm_report in norm_name:
            if len(norm_report) > best_len:
                best_match = report_path
                best_len = len(norm_report)

    # Only accept fuzzy match if it covers most of the name (>90%)
    if best_match and best_len >= 0.9 * max(len(norm_name), len(normalise(best_match))):
        return best_match

    return None


def load_mt5_metrics(mt5_folder, strategy_names):
    mt5_reports = find_mt5_reports(mt5_folder)
    if not mt5_reports:
        print(f"  No MT5 HTML reports found in '{mt5_folder}'")
        return {}
    print(f"  Found {len(mt5_reports)} MT5 report file(s)")

    results = {}
    for name in strategy_names:
        report_path = match_strategy_to_report(name, mt5_reports)
        if report_path:
            metrics = parse_mt5_report(report_path)
            if metrics:
                results[name] = metrics
                print(f"    + Matched: {name} -> {os.path.basename(report_path)}")
            else:
                print(f"    x Failed to parse: {os.path.basename(report_path)}")
        else:
            print(f"    - No MT5 report found for: {name}")
    return results

# ============================================================================
# MT5 RANKING LOGIC
# ============================================================================
def compute_mt5_rankings(mt5_metrics):
    if not mt5_metrics:
        return pd.DataFrame()

    rows = []
    for name, m in mt5_metrics.items():
        row = {'Strategy': name}
        row.update(m)
        rows.append(row)
    df = pd.DataFrame(rows)

    total_weight = sum(MT5_HIGHER_IS_BETTER.values()) + sum(MT5_LOWER_IS_BETTER.values())
    if not np.isclose(total_weight, 1.0):
        print(f"  WARNING: MT5 ranking weights sum to {total_weight:.4f}, not 1.0")

    available_higher = {k: v for k, v in MT5_HIGHER_IS_BETTER.items() if k in df.columns}
    available_lower = {k: v for k, v in MT5_LOWER_IS_BETTER.items() if k in df.columns}

    missing = (set(MT5_HIGHER_IS_BETTER) | set(MT5_LOWER_IS_BETTER)) - set(df.columns)
    if missing:
        print(f"  Note: Some ranking metrics not available: {missing}")

    avail_total = sum(available_higher.values()) + sum(available_lower.values())
    if avail_total == 0:
        print("  ERROR: No ranking metrics available")
        return pd.DataFrame()

    for col in available_higher:
        df[f'_pctrank_{col}'] = df[col].rank(pct=True).fillna(0)
    for col in available_lower:
        df[f'_pctrank_{col}'] = (1 - df[col].rank(pct=True)).fillna(0)

    df['Composite Score'] = sum(
        df[f'_pctrank_{col}'] * (weight / avail_total)
        for col, weight in {**available_higher, **available_lower}.items()
    )

    df['Rank'] = df['Composite Score'].rank(ascending=False, na_option='bottom').astype(int)
    df = df.sort_values('Rank').reset_index(drop=True)
    df = df.drop(columns=[c for c in df.columns if c.startswith('_pctrank_')])
    return df


# ============================================================================
# ANALYSIS FUNCTIONS
# ============================================================================
def build_pnl_series(strategies, freq='D'):
    pnl_dict = {}
    for name, df in strategies.items():
        if freq == 'D':
            key = df['Close time'].dt.date
        elif freq == 'W':
            key = df['Close time'].dt.to_period('W').apply(lambda x: x.start_time.date())
        elif freq == 'M':
            key = df['Close time'].dt.to_period('M').apply(lambda x: x.start_time.date())
        else:
            key = df['Close time'].dt.date
        pnl_dict[name] = df.groupby(key)['Profit/Loss'].sum()
    # Use NaN for periods with no trades (not 0) so we can distinguish
    # "no activity" from "traded and broke even"
    return pd.DataFrame(pnl_dict)


def compute_pairwise_correlation(pnl_df, min_observations=10):
    """
    Compute pairwise correlation excluding periods where BOTH strategies
    had no activity (NaN). This prevents shared quiet periods (holidays,
    low-volatility weeks) from inflating correlation artificially.

    For each pair, periods are included if at least one strategy traded.
    NaN values (no trades) are treated as 0 only when the other strategy
    was active during that period.

    Args:
        pnl_df: DataFrame with NaN for no-trade periods (from build_pnl_series)
        min_observations: minimum shared active periods required for a valid
                         correlation (returns 0 if insufficient data)

    Returns:
        Correlation matrix (DataFrame)
    """
    names = pnl_df.columns.tolist()
    n = len(names)
    corr_matrix = pd.DataFrame(1.0, index=names, columns=names)

    for i in range(n):
        for j in range(i + 1, n):
            s1 = pnl_df[names[i]]
            s2 = pnl_df[names[j]]

            # Mask: at least one strategy had activity (non-NaN)
            active_mask = s1.notna() | s2.notna()
            s1_active = s1[active_mask].fillna(0)
            s2_active = s2[active_mask].fillna(0)

            if len(s1_active) >= min_observations:
                corr_val = s1_active.corr(s2_active)
                corr_val = 0.0 if pd.isna(corr_val) else corr_val
            else:
                corr_val = 0.0

            corr_matrix.loc[names[i], names[j]] = round(corr_val, 6)
            corr_matrix.loc[names[j], names[i]] = round(corr_val, 6)

    return corr_matrix


def compute_trade_overlap(df1, df2):
    open1 = df1['Open time'].values
    close1 = df1['Close time'].values
    open2 = df2['Open time'].values
    close2 = df2['Close time'].values
    types1 = df1['Type'].values
    types2 = df2['Type'].values
    long_types = {'Buy', 'BuyStop', 'BuyLimit'}
    overlaps = 0
    same_dir = 0
    opp_dir = 0
    for i in range(len(df1)):
        mask = (open2 < close1[i]) & (close2 > open1[i])
        if mask.any():
            overlaps += 1
            dir1_long = types1[i] in long_types
            for j in np.where(mask)[0]:
                dir2_long = types2[j] in long_types
                if dir1_long == dir2_long:
                    same_dir += 1
                else:
                    opp_dir += 1
    return overlaps, same_dir, opp_dir


def compute_drawdown_info(df):
    df_sorted = df.sort_values('Close time').reset_index(drop=True)
    equity = df_sorted['Profit/Loss'].cumsum()
    peak = equity.cummax()
    dd = equity - peak
    max_dd_idx = dd.idxmin()
    max_dd_val = dd.iloc[max_dd_idx]
    max_dd_end = df_sorted.loc[max_dd_idx, 'Close time']
    peak_idx = equity.iloc[:max_dd_idx + 1].idxmax()
    max_dd_start = df_sorted.loc[peak_idx, 'Close time']
    duration_days = (max_dd_end - max_dd_start).days
    return max_dd_val, max_dd_start, max_dd_end, duration_days


def compute_strategy_stats(strategies):
    stats = {}
    long_types = {'Buy', 'BuyStop', 'BuyLimit'}
    short_types = {'Sell', 'SellStop', 'SellLimit'}
    for name, df in strategies.items():
        n_long = df['Type'].isin(long_types).sum()
        n_short = df['Type'].isin(short_types).sum()
        win_rate = (df['Profit/Loss'] > 0).mean()
        avg_trade = df['Profit/Loss'].mean()
        total_pnl = df['Profit/Loss'].sum()
        max_dd, dd_start, dd_end, dd_days = compute_drawdown_info(df)
        direction = 'Long Only' if n_short == 0 else ('Short Only' if n_long == 0 else 'Both')
        stats[name] = {
            'trades': len(df), 'long': n_long, 'short': n_short,
            'direction': direction, 'win_rate': win_rate,
            'avg_trade': avg_trade, 'total_pnl': total_pnl,
            'max_dd': max_dd, 'dd_start': dd_start,
            'dd_end': dd_end, 'dd_days': dd_days,
        }
    return stats


def identify_clusters(names, corr_matrix, threshold=0.5):
    clusters = []
    assigned = set()
    for name in names:
        if name in assigned:
            continue
        cluster = [name]
        assigned.add(name)
        for other in names:
            if other in assigned:
                continue
            if abs(corr_matrix.loc[name, other]) >= threshold:
                cluster.append(other)
                assigned.add(other)
        clusters.append(cluster)
    return clusters

# ============================================================================
# EXCEL REPORT GENERATION — Styles
# ============================================================================
HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
SUBHEADER_FONT = Font(name='Arial', bold=True, size=10)
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
NORMAL_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='2F5496')
WARNING_FONT = Font(name='Arial', bold=True, size=10, color='CC0000')
SUCCESS_FONT = Font(name='Arial', bold=True, size=10, color='006600')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
FILL_LOW = PatternFill('solid', fgColor='6BCB77')
FILL_MODERATE = PatternFill('solid', fgColor='B5E48C')
FILL_HIGH = PatternFill('solid', fgColor='FFD93D')
FILL_VERY_HIGH = PatternFill('solid', fgColor='FF6B6B')
FILL_SELF = PatternFill('solid', fgColor='D9D9D9')
KEEP_FILL = PatternFill('solid', fgColor='C6EFCE')
KEEP_FONT = Font(name='Arial', bold=True, size=10, color='006100')
ABANDON_FILL = PatternFill('solid', fgColor='FFC7CE')
ABANDON_FONT = Font(name='Arial', size=10, color='9C0006')
SUBTITLE_FONT = Font(name='Arial', size=10, italic=True)
TOP5_FILL = PatternFill('solid', fgColor='C6EFCE')
TOP10_FILL = PatternFill('solid', fgColor='D9E2F3')
TOP25_FILL = PatternFill('solid', fgColor='FFF2CC')
SECTION_FONT = Font(name='Arial', bold=True, size=11, color='2F5496')
BODY_FONT = Font(name='Arial', size=10)
DATA_FONT = Font(name='Arial', size=10)
DATA_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')


def style_header_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = THIN_BORDER
    return cell


def style_cell(ws, row, col, value, fmt=None, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BOLD_FONT if bold else NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if fmt:
        cell.number_format = fmt
    return cell


def color_corr_cell(ws, row, col, value):
    cell = style_cell(ws, row, col, value, fmt='0.000')
    if value is not None and not np.isnan(value):
        if abs(value) >= 1.0 - 1e-9:
            cell.fill = FILL_SELF
        elif abs(value) >= CORR_VERY_HIGH:
            cell.fill = FILL_VERY_HIGH
        elif abs(value) >= CORR_HIGH:
            cell.fill = FILL_HIGH
        elif abs(value) >= CORR_MODERATE:
            cell.fill = FILL_MODERATE
        else:
            cell.fill = FILL_LOW
    return cell


def write_title(ws, row, text, merge_end_col=None):
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_end_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='left')
    return cell


def write_colour_key(ws, row):
    ws.cell(row=row, column=1, value='Colour Key:').font = BOLD_FONT
    keys = [
        (2, '< 0.3 Low', '6BCB77'),
        (3, '0.3-0.5 Moderate', 'B5E48C'),
        (4, '0.5-0.7 High', 'FFD93D'),
        (5, '> 0.7 Very High', 'FF6B6B'),
    ]
    for col, label, color in keys:
        c = ws.cell(row=row, column=col, value=label)
        c.fill = PatternFill('solid', fgColor=color)
        c.font = Font(name='Arial', bold=True, size=9)
        c.alignment = Alignment(horizontal='center')
        c.border = THIN_BORDER


def write_correlation_sheet(wb, sheet_name, tab_color, title, names, corr_matrix):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color
    n = len(names)
    write_title(ws, 1, title, merge_end_col=n + 1)
    row = 3
    style_header_cell(ws, row, 1, '')
    for c, name in enumerate(names, 2):
        style_header_cell(ws, row, c, name)
    for i, n1 in enumerate(names):
        r = row + 1 + i
        cell = style_cell(ws, r, 1, n1, bold=True)
        cell.fill = SUBHEADER_FILL
        for j, n2 in enumerate(names):
            color_corr_cell(ws, r, j + 2, corr_matrix.loc[n1, n2])
    write_colour_key(ws, row + n + 2)
    ws.column_dimensions['A'].width = max(len(n) for n in names) + 4
    for col in range(2, n + 2):
        ws.column_dimensions[get_column_letter(col)].width = max(len(n) for n in names) + 2

# ============================================================================
# MT5 RANKING SHEET GENERATION
# ============================================================================
def write_mt5_ranking_sheet(wb, ranked_df, keep_strategies):
    if ranked_df.empty:
        return
    ws = wb.create_sheet('MT5 Backtest Ranking')
    ws.sheet_properties.tabColor = '7030A0'

    ws.merge_cells('A1:Q1')
    ws.cell(row=1, column=1,
            value='MT5 Backtest Rankings \u2014 All Strategies Passing Monte Carlo').font = TITLE_FONT
    ws.merge_cells('A2:Q2')
    ws.cell(row=2, column=1,
            value=f'{len(ranked_df)} strategies ranked by multi-metric composite score '
                  f'(KEEP strategies shown in bold)').font = SUBTITLE_FONT

    output_columns = [
        ('Rank',                'Rank',                 6,  '#,##0'),
        ('Strategy',            'Strategy',             30, None),
        ('Composite Score',     'Composite\nScore',     11, '0.000'),
        ('Symbol',              'Symbol',               16, None),
        ('Total Net Profit',    'Net Profit',           13, '$#,##0.00'),
        ('Ret/DD Ratio',        'Ret/DD\nRatio',        10, '0.00'),
        ('MC95 Ret/DD',         'MC95\nRet/DD',         10, '0.00'),
        ('Win/Loss Ratio',      'Win/Loss\nRatio',      10, '0.00'),
        ('Profit Factor',       'Profit\nFactor',       10, '0.00'),
        ('Sharpe Ratio',        'Sharpe\nRatio',        10, '0.00'),
        ('Recovery Factor',     'Recovery\nFactor',     10, '0.00'),
        ('LR Correlation',      'LR Corr.\n(Stability)', 11, '0.00'),
        ('Win Rate %',          'Win Rate\n%',          9,  '0.0'),
        ('Total Trades',        '# Trades',             9,  '#,##0'),
        ('Expected Payoff',     'Expected\nPayoff',     10, '0.00'),
        ('Balance DD Max $',    'Max DD\n($)',          12, '$#,##0.00'),
        ('Balance DD Rel %',    'Max DD\n(%)',          9,  '0.0'),
        ('LR Standard Error',   'LR Std\nError',       10, '0.00'),
    ]

    valid_columns = [(col, hdr, w, fmt) for col, hdr, w, fmt in output_columns if col in ranked_df.columns]

    header_row = 4
    for col_idx, (_, header, width, _) in enumerate(valid_columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 35

    for row_idx, (_, row) in enumerate(ranked_df.iterrows(), header_row + 1):
        rank = row.get('Rank', row_idx - header_row)
        strategy_name = row.get('Strategy', '')
        for col_idx, (col_name, _, _, fmt) in enumerate(valid_columns, 1):
            val = row.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col_name == 'Strategy':
                cell.alignment = LEFT_ALIGN
            else:
                cell.alignment = DATA_ALIGN
            if fmt:
                cell.number_format = fmt
            if isinstance(rank, (int, float)) and not np.isnan(rank):
                if rank <= 5:
                    cell.fill = TOP5_FILL
                elif rank <= 10:
                    cell.fill = TOP10_FILL
                elif rank <= 25:
                    cell.fill = TOP25_FILL
        if strategy_name in keep_strategies:
            for col_idx in range(1, len(valid_columns) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.font = Font(name='Arial', bold=True, size=10)

    last_col = get_column_letter(len(valid_columns))
    last_row = header_row + len(ranked_df)
    ws.auto_filter.ref = f'A{header_row}:{last_col}{last_row}'
    ws.freeze_panes = f'A{header_row + 1}'
    return ws


def write_mt5_methodology_sheet(wb):
    ws = wb.create_sheet('MT5 Ranking Methodology')
    ws.sheet_properties.tabColor = '7030A0'
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 70

    r = 1
    ws.cell(row=r, column=2, value='MT5 Backtest Ranking Methodology').font = TITLE_FONT
    r += 2
    ws.cell(row=r, column=2, value='Overview').font = SECTION_FONT
    r += 1
    for line in [
        'Strategies that pass Monte Carlo testing in StrategyQuant X are validated',
        'by running backtests in MetaTrader 5. The MT5 HTML reports are then parsed',
        'to extract performance metrics, which are used to rank strategies using the',
        'same percentile-based composite scoring methodology as the SQX databank ranker.',
    ]:
        ws.cell(row=r, column=2, value=line).font = BODY_FONT
        r += 1
    r += 1

    ws.cell(row=r, column=2, value='Scoring Method').font = SECTION_FONT
    r += 1
    for line in [
        'Each metric is converted to a percentile rank (0-1), then weighted and summed.',
        'Metrics where lower is better (e.g. drawdown %) are inverted before weighting.',
        'This ensures strategies must be consistently good across multiple dimensions.',
    ]:
        ws.cell(row=r, column=2, value=line).font = BODY_FONT
        r += 1
    r += 1

    ws.cell(row=r, column=2, value='Metric').font = BOLD_FONT
    ws.cell(row=r, column=3, value='Weight').font = BOLD_FONT
    ws.cell(row=r, column=4, value='Rationale').font = BOLD_FONT
    for c in [2, 3, 4]:
        ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor='D9E2F3')
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 1

    methodology_rows = [
        ('Net Profit',       '15%', 'Absolute profitability \u2014 ensures the strategy generates meaningful returns'),
        ('Ret/DD Ratio',     '15%', 'Net profit / max equity drawdown \u2014 risk-adjusted return quality'),
        ('Profit Factor',    '10%', 'Gross profit / gross loss \u2014 measures overall reward-to-risk'),
        ('Sharpe Ratio',     '10%', 'Risk-adjusted return considering volatility \u2014 penalises inconsistent curves'),
        ('Recovery Factor',  '10%', 'Net profit / max drawdown \u2014 how quickly the strategy recovers from losses'),
        ('LR Correlation',   '10%', 'Linear regression R of equity curve \u2014 measures equity curve smoothness/stability'),
        ('Win/Loss Ratio',   '5%',  'Average win / average loss \u2014 reward-to-risk per trade'),
        ('# Trades',         '5%',  'Statistical significance \u2014 more trades = more confidence in the edge'),
        ('Expected Payoff',  '5%',  'Average P&L per trade \u2014 how much edge each trade provides'),
        ('Win Rate %',       '5%',  'Percentage of winning trades \u2014 consistency of the edge'),
        ('Max DD %',         '5%',  'Inverted \u2014 lower max drawdown percentage is better'),
        ('LR Std Error',     '5%',  'Inverted \u2014 lower standard error = smoother equity curve'),
    ]
    for metric, weight, rationale in methodology_rows:
        ws.cell(row=r, column=2, value=metric).font = BODY_FONT
        ws.cell(row=r, column=3, value=weight).font = BODY_FONT
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4, value=rationale).font = BODY_FONT
        for c in [2, 3, 4]:
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=2, value='Colour Coding').font = SECTION_FONT
    r += 1
    for label, fill in [('Top 5', TOP5_FILL), ('Top 10', TOP10_FILL), ('Top 25', TOP25_FILL)]:
        ws.cell(row=r, column=2, value=label).font = BODY_FONT
        ws.cell(row=r, column=3).fill = fill
        ws.cell(row=r, column=3).border = THIN_BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=2, value='Data Source').font = SECTION_FONT
    r += 1
    for note in [
        '\u2022 Metrics are parsed directly from MT5 Strategy Tester HTML reports',
        '\u2022 Reports must be in the same folder as the trade CSV files (or specified via --mt5-reports)',
        '\u2022 Strategy names are matched by normalising filenames (underscores / spaces / dots)',
        '\u2022 Ret/DD Ratio and Win/Loss Ratio are computed from the parsed MT5 values',
        '\u2022 Weights are adjustable in the MT5_HIGHER_IS_BETTER / MT5_LOWER_IS_BETTER config at top of script',
    ]:
        ws.cell(row=r, column=2, value=note).font = BODY_FONT
        r += 1

# ============================================================================
# MAIN REPORT GENERATION
# ============================================================================
def generate_report(folder_path, strategies, stats, names,
                    corr_daily, corr_weekly, corr_monthly,
                    overlap_data, clusters, mt5_folder=None,
                    mc95_data=None, mc_failed=None):
    wb = Workbook()
    n = len(names)

    # Sheet 1: Summary
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_properties.tabColor = '2F5496'
    write_title(ws, 1, 'Strategy Correlation Analysis', merge_end_col=8)
    ws.cell(row=2, column=1, value=f'{n} strategies from: {folder_path}').font = Font(name='Arial', size=9, italic=True)

    row = 4
    headers = ['Strategy', 'Trades', 'Long', 'Short', 'Direction', 'Win Rate', 'Avg Trade ($)', 'Total P&L ($)']
    for c, h in enumerate(headers, 1):
        style_header_cell(ws, row, c, h)

    for i, name in enumerate(names):
        s = stats[name]
        r = row + 1 + i
        style_cell(ws, r, 1, name, bold=True)
        style_cell(ws, r, 2, s['trades'], fmt='#,##0')
        style_cell(ws, r, 3, s['long'], fmt='#,##0')
        style_cell(ws, r, 4, s['short'], fmt='#,##0')
        dir_cell = style_cell(ws, r, 5, s['direction'])
        if s['direction'] == 'Long Only':
            dir_cell.font = Font(name='Arial', size=10, color='0066CC')
        elif s['direction'] == 'Short Only':
            dir_cell.font = Font(name='Arial', size=10, color='CC0000')
        style_cell(ws, r, 6, s['win_rate'], fmt='0.0%')
        style_cell(ws, r, 7, s['avg_trade'], fmt='#,##0.00')
        style_cell(ws, r, 8, s['total_pnl'], fmt='#,##0.00')

    long_only = [nm for nm in names if stats[nm]['direction'] == 'Long Only']
    both_dir = [nm for nm in names if stats[nm]['direction'] == 'Both']
    short_only = [nm for nm in names if stats[nm]['direction'] == 'Short Only']

    info_row = row + n + 2
    if long_only:
        cell = ws.cell(row=info_row, column=1,
                       value=f'Direction Bias: {len(long_only)}/{n} strategies are Long Only, '
                             f'{len(both_dir)} trade Both directions, {len(short_only)} Short Only.')
        cell.font = WARNING_FONT if len(long_only) > n * 0.6 else BOLD_FONT
        ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=8)
    write_colour_key(ws, info_row + 2)
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['A'].width = max(len(nm) for nm in names) + 4

    # Correlation sheets
    write_correlation_sheet(wb, 'Daily Correlation', '4472C4', 'Daily P&L Correlation Matrix', names, corr_daily)
    write_correlation_sheet(wb, 'Weekly Correlation', '4472C4', 'Weekly P&L Correlation Matrix', names, corr_weekly)
    write_correlation_sheet(wb, 'Monthly Correlation', '4472C4', 'Monthly P&L Correlation Matrix', names, corr_monthly)

    # Trade Overlap
    ws5 = wb.create_sheet('Trade Overlap')
    ws5.sheet_properties.tabColor = 'ED7D31'
    write_title(ws5, 1, 'Pairwise Trade Overlap Analysis', merge_end_col=6)
    row = 3
    for c, h in enumerate(['Pair', 'Overlapping Trades', 'Overlap %', 'Same Direction', 'Opp Direction', 'Assessment'], 1):
        style_header_cell(ws5, row, c, h)
    r = row + 1
    for (n1, n2), data in sorted(overlap_data.items(), key=lambda x: x[1]['pct'], reverse=True):
        style_cell(ws5, r, 1, f'{n1} vs {n2}', bold=True)
        style_cell(ws5, r, 2, f"{data['overlaps']}/{data['total']}")
        style_cell(ws5, r, 3, data['pct'], fmt='0.0%')
        style_cell(ws5, r, 4, data['same'], fmt='#,##0')
        style_cell(ws5, r, 5, data['opp'], fmt='#,##0')
        pct = data['pct']
        if pct >= OVERLAP_VERY_HIGH:
            assessment, color = 'Very High - Likely Redundant', 'FF6B6B'
        elif pct >= OVERLAP_HIGH:
            assessment, color = 'High - Significant Overlap', 'FFD93D'
        elif pct >= OVERLAP_MODERATE:
            assessment, color = 'Moderate', 'B5E48C'
        else:
            assessment, color = 'Low - Good Diversification', '6BCB77'
        c = style_cell(ws5, r, 6, assessment)
        c.fill = PatternFill('solid', fgColor=color)
        r += 1
    for col in range(1, 7):
        ws5.column_dimensions[get_column_letter(col)].width = 24
    ws5.column_dimensions['A'].width = max(len(nm) for nm in names) * 2 + 8

    # Overlap Matrix
    ws5b = wb.create_sheet('Overlap Matrix')
    ws5b.sheet_properties.tabColor = 'ED7D31'
    write_title(ws5b, 1, 'Trade Overlap % Matrix (row strategy overlap with column strategy)', merge_end_col=n + 1)
    row = 3
    style_header_cell(ws5b, row, 1, '')
    for c, name in enumerate(names, 2):
        style_header_cell(ws5b, row, c, name)
    for i, n1 in enumerate(names):
        r = row + 1 + i
        cell = style_cell(ws5b, r, 1, n1, bold=True)
        cell.fill = SUBHEADER_FILL
        for j, n2 in enumerate(names):
            if n1 == n2:
                c = style_cell(ws5b, r, j + 2, '-')
                c.fill = FILL_SELF
            else:
                key = (n1, n2) if (n1, n2) in overlap_data else (n2, n1)
                if key in overlap_data:
                    if key == (n1, n2):
                        pct = overlap_data[key]['pct']
                    else:
                        pct = overlap_data[key]['overlaps'] / len(strategies[n1]) if len(strategies[n1]) > 0 else 0
                    c = style_cell(ws5b, r, j + 2, pct, fmt='0.0%')
                    if pct >= OVERLAP_VERY_HIGH: c.fill = FILL_VERY_HIGH
                    elif pct >= OVERLAP_HIGH: c.fill = FILL_HIGH
                    elif pct >= OVERLAP_MODERATE: c.fill = FILL_MODERATE
                    else: c.fill = FILL_LOW
    ws5b.column_dimensions['A'].width = max(len(nm) for nm in names) + 4
    for col in range(2, n + 2):
        ws5b.column_dimensions[get_column_letter(col)].width = max(len(nm) for nm in names) + 2

    # Drawdown Periods
    ws6 = wb.create_sheet('Drawdown Periods')
    ws6.sheet_properties.tabColor = 'A5A5A5'
    write_title(ws6, 1, 'Max Drawdown Period Analysis', merge_end_col=5)
    row = 3
    for c, h in enumerate(['Strategy', 'Max Drawdown ($)', 'DD Start', 'DD End', 'Duration (days)'], 1):
        style_header_cell(ws6, row, c, h)
    for i, name in enumerate(names):
        s = stats[name]
        r = row + 1 + i
        style_cell(ws6, r, 1, name, bold=True)
        style_cell(ws6, r, 2, s['max_dd'], fmt='#,##0.00')
        style_cell(ws6, r, 3, s['dd_start'].strftime('%Y-%m-%d'))
        style_cell(ws6, r, 4, s['dd_end'].strftime('%Y-%m-%d'))
        style_cell(ws6, r, 5, s['dd_days'], fmt='#,##0')
    for col in range(1, 6):
        ws6.column_dimensions[get_column_letter(col)].width = 20
    ws6.column_dimensions['A'].width = max(len(nm) for nm in names) + 4

    # Clusters
    ws7 = wb.create_sheet('Correlation Clusters')
    ws7.sheet_properties.tabColor = '70AD47'
    write_title(ws7, 1, 'Strategy Clusters (Weekly Correlation >= 0.5)', merge_end_col=3)
    ws7.cell(row=2, column=1,
             value='Strategies within the same cluster are correlated \u2014 consider picking the best performer from each.').font = \
        Font(name='Arial', size=10, italic=True)
    row = 4
    for c, h in enumerate(['Cluster', 'Strategies', 'Count'], 1):
        style_header_cell(ws7, row, c, h)
    for i, cluster in enumerate(clusters):
        r = row + 1 + i
        style_cell(ws7, r, 1, f'Cluster {i + 1}', bold=True)
        style_cell(ws7, r, 2, ', '.join(cluster))
        cell = style_cell(ws7, r, 3, len(cluster), fmt='#,##0')
        if len(cluster) == 1: cell.fill = FILL_LOW
        elif len(cluster) <= 3: cell.fill = FILL_MODERATE
        else: cell.fill = FILL_HIGH
    ws7.column_dimensions['A'].width = 14
    ws7.column_dimensions['B'].width = max(40, max(len(', '.join(c)) for c in clusters) + 4)
    ws7.column_dimensions['C'].width = 10
    summary_row = row + len(clusters) + 2
    ws7.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
    ws7.cell(row=summary_row, column=1,
             value=f'{len(clusters)} independent cluster(s) identified from {n} strategies. '
                   f'Consider selecting ~{len(clusters)} strategies (1 per cluster) for optimal diversification.').font = SUCCESS_FONT

    # Best Pairs
    ws8 = wb.create_sheet('Best Pairs')
    ws8.sheet_properties.tabColor = '70AD47'
    write_title(ws8, 1, 'Strategy Pairs Ranked by Diversification Benefit (Lowest Weekly Correlation)', merge_end_col=4)
    row = 3
    for c, h in enumerate(['Rank', 'Pair', 'Weekly Correlation', 'Daily Correlation'], 1):
        style_header_cell(ws8, row, c, h)
    pairs = []
    for i, n1 in enumerate(names):
        for j, n2 in enumerate(names):
            if j > i:
                pairs.append((n1, n2, corr_weekly.loc[n1, n2], corr_daily.loc[n1, n2]))
    pairs.sort(key=lambda x: abs(x[2]))
    for rank, (n1, n2, wcorr, dcorr) in enumerate(pairs, 1):
        r = row + rank
        style_cell(ws8, r, 1, rank)
        style_cell(ws8, r, 2, f'{n1} vs {n2}', bold=True)
        color_corr_cell(ws8, r, 3, wcorr)
        color_corr_cell(ws8, r, 4, dcorr)
    ws8.column_dimensions['A'].width = 8
    ws8.column_dimensions['B'].width = max(len(nm) for nm in names) * 2 + 8
    ws8.column_dimensions['C'].width = 20
    ws8.column_dimensions['D'].width = 20

    # =========================================================================
    # STEP 1: Rank ALL strategies on MT5 metrics FIRST (before cluster selection)
    # =========================================================================
    ranked_df = None
    mt5_metrics = None
    strategy_scores = {}  # name -> composite score (for cluster selection)

    if mt5_folder:
        print("\nLoading MT5 backtest reports for ALL strategies...")
        mt5_metrics = load_mt5_metrics(mt5_folder, names)
        if mt5_metrics:
            print(f"\nRanking {len(mt5_metrics)} strategies from MT5 reports...")
            ranked_df = compute_mt5_rankings(mt5_metrics)
            if not ranked_df.empty:
                # Add MC95 Ret/DD column if MC data available
                if mc95_data:
                    mc95_vals = []
                    for _, row in ranked_df.iterrows():
                        rname = row.get('Strategy', '')
                        matched = match_csv_to_mc(rname, set(mc95_data.keys()))
                        if matched and matched in mc95_data:
                            mc95_vals.append(mc95_data[matched].get('mc95_ret_dd', None))
                        else:
                            mc95_vals.append(None)
                    ranked_df['MC95 Ret/DD'] = mc95_vals

                # Build score lookup for cluster selection
                for _, row in ranked_df.iterrows():
                    strategy_scores[row['Strategy']] = row['Composite Score']

    # =========================================================================
    # STEP 2: Portfolio Selection — use composite score to pick cluster survivors
    # =========================================================================
    ws_ps = wb.create_sheet('Portfolio Selection', 0)
    ws_ps.sheet_properties.tabColor = '70AD47'

    keep = {}
    abandon = {}
    for cid, members in enumerate(clusters, 1):
        if len(members) == 1:
            keep[members[0]] = cid
        else:
            if strategy_scores:
                # Use composite score (higher is better) — the key improvement
                best = max(members, key=lambda s: strategy_scores.get(s, -1))
            else:
                # Fallback to P&L if no MT5 reports available
                best = max(members, key=lambda s: stats[s]['total_pnl'])
            keep[best] = cid
            for s in members:
                if s != best:
                    abandon[s] = (cid, best)

    max_name_len = max(len(nm) for nm in names)
    ws_ps.merge_cells('A1:H1')
    ws_ps.cell(row=1, column=1, value='Portfolio Selection \u2014 Which Strategies to Keep vs Abandon').font = TITLE_FONT
    selection_method = 'composite score' if strategy_scores else 'total P&L'
    ws_ps.merge_cells('A2:H2')
    ws_ps.cell(row=2, column=1,
               value=f'Selecting best performer (by {selection_method}) from each correlation cluster: '
                     f'{len(keep)} KEEP, {len(abandon)} ABANDON').font = SUBTITLE_FONT

    row = 4
    ps_headers = ['Strategy', 'Decision', 'Cluster', 'Reason', 'Direction', 'Total P&L ($)', 'Avg Trade ($)', 'Win Rate']
    for c, h in enumerate(ps_headers, 1):
        style_header_cell(ws_ps, row, c, h)

    r = row + 1
    for name in sorted(keep.keys(), key=lambda s: keep[s]):
        cid = keep[name]
        s = stats[name]
        cluster_size = len(clusters[cid - 1])
        if cluster_size == 1:
            reason = 'Only strategy in cluster'
        elif strategy_scores:
            score = strategy_scores.get(name, 0)
            reason = f'Best composite score ({score:.3f}) in cluster of {cluster_size}'
        else:
            reason = f'Best P&L in cluster of {cluster_size}'
        for col_idx, val in enumerate([name, '\u2713 KEEP', f'Cluster {cid}', reason,
                                       s['direction'], s['total_pnl'], s['avg_trade'], s['win_rate']], 1):
            fmt = None
            if col_idx == 6: fmt = '#,##0.00'
            elif col_idx == 7: fmt = '#,##0.00'
            elif col_idx == 8: fmt = '0.0%'
            cell = style_cell(ws_ps, r, col_idx, val, fmt=fmt)
            cell.font = KEEP_FONT
            cell.fill = KEEP_FILL
        r += 1

    r += 1
    for name in sorted(abandon.keys(), key=lambda s: abandon[s][0]):
        cid, replaced_by = abandon[name]
        s = stats[name]
        if strategy_scores:
            my_score = strategy_scores.get(name, 0)
            their_score = strategy_scores.get(replaced_by, 0)
            reason = f'Redundant \u2014 correlated with {replaced_by} (score {their_score:.3f} vs {my_score:.3f})'
        else:
            reason = f'Redundant \u2014 correlated with {replaced_by}'
        for col_idx, val in enumerate([name, '\u2717 ABANDON', f'Cluster {cid}', reason,
                                       s['direction'], s['total_pnl'], s['avg_trade'], s['win_rate']], 1):
            fmt = None
            if col_idx == 6: fmt = '#,##0.00'
            elif col_idx == 7: fmt = '#,##0.00'
            elif col_idx == 8: fmt = '0.0%'
            cell = style_cell(ws_ps, r, col_idx, val, fmt=fmt)
            cell.font = ABANDON_FONT
            cell.fill = ABANDON_FILL
        r += 1

    r += 2
    ws_ps.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws_ps.cell(row=r, column=1, value='Summary').font = Font(name='Arial', bold=True, size=12, color='2F5496')
    r += 1
    keep_pnl = sum(stats[s]['total_pnl'] for s in keep)
    abandon_pnl = sum(stats[s]['total_pnl'] for s in abandon)
    keep_long_only = sum(1 for s in keep if stats[s]['direction'] == 'Long Only')
    keep_both = sum(1 for s in keep if stats[s]['direction'] == 'Both')
    summaries = [
        f'Portfolio reduced from {n} \u2192 {len(keep)} strategies (removed {len(abandon)} redundant)',
        f'KEEP total P&L: ${keep_pnl:,.0f}  |  ABANDON total P&L: ${abandon_pnl:,.0f}',
        f'Direction mix in KEEP portfolio: {keep_long_only} Long Only + {keep_both} Both Directions',
        f'Selection method: Highest {selection_method} from each weekly correlation cluster (threshold \u2265 0.5)',
    ]
    for line in summaries:
        ws_ps.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws_ps.cell(row=r, column=1, value=line).font = NORMAL_FONT
        r += 1
    ps_widths = [max_name_len + 4, 14, 12, max_name_len + 24, 12, 14, 14, 10]
    for i, w in enumerate(ps_widths, 1):
        ws_ps.column_dimensions[get_column_letter(i)].width = w

    # Write MT5 Ranking sheets (data already computed in Step 1)
    if ranked_df is not None and not ranked_df.empty:
        write_mt5_ranking_sheet(wb, ranked_df, set(keep.keys()))
        write_mt5_methodology_sheet(wb)
        print(f"\n{'='*90}")
        print(f"  MT5 BACKTEST RANKINGS (ALL strategies)")
        print(f"{'='*90}")
        print(f"{'Rank':>4}  {'Score':>6}  {'Strategy':<32}  {'Profit':>10}  {'Ret/DD':>6}  "
              f"{'PF':>5}  {'Sharpe':>6}  {'LR Corr':>7}  {'Trades':>6}  {'Status':>8}")
        print(f"{'-'*100}")
        for _, row in ranked_df.iterrows():
            sname = row['Strategy']
            status = 'KEEP' if sname in keep else 'ABANDON'
            print(f"#{int(row['Rank']):3d}  {row['Composite Score']:6.3f}  "
                  f"{sname:<32.32s}  "
                  f"${row.get('Total Net Profit', 0):9.2f}  "
                  f"{row.get('Ret/DD Ratio', 0):6.2f}  "
                  f"{row.get('Profit Factor', 0):5.2f}  "
                  f"{row.get('Sharpe Ratio', 0):6.2f}  "
                  f"{row.get('LR Correlation', 0):7.2f}  "
                  f"{int(row.get('Total Trades', 0)):6d}  "
                  f"{'  ' + status:>8s}")
        print(f"{'='*100}")
    elif mt5_folder:
        print("  No MT5 reports matched strategies - skipping ranking sheets.")

    output_path = os.path.join(folder_path, 'Strategy_Correlation_Analysis.xlsx')
    wb.save(output_path)

    # Generate HTML Dashboard
    print("\nGenerating equity comparison charts...")
    try:
        chart_map = generate_equity_charts(folder_path, strategies, names, mt5_folder=mt5_folder)
        print(f"  Generated {len(chart_map)} equity chart(s)")
    except Exception as e:
        import traceback
        print(f"  WARNING: Equity chart generation failed: {e}")
        traceback.print_exc()
        chart_map = {}

    print("\nLoading MT5 full overviews for dashboard...")
    mt5_overviews = {}
    if mt5_folder:
        try:
            mt5_overviews = load_mt5_overviews(mt5_folder, names)
            print(f"  Loaded {len(mt5_overviews)} MT5 overviews")
        except Exception as e:
            print(f"  WARNING: Overview loading failed: {e}")

    print("\nLoading strategy pseudo code files...")
    strategy_codes = {}
    try:
        strategy_codes = load_strategy_pseudo_codes(folder_path, names)
    except Exception as e:
        print(f"  WARNING: Pseudo code loading failed: {e}")

    print("\nLoading SQX metadata (Complexity etc.)...")
    sqx_metadata = {}
    try:
        sqx_metadata = load_sqx_metadata(folder_path, names)
    except Exception as e:
        print(f"  WARNING: SQX metadata loading failed: {e}")

    print("\nGenerating HTML dashboard...")
    try:
        dashboard_path = generate_dashboard(
            folder_path, strategies, stats, names,
            corr_daily, corr_weekly, corr_monthly,
            overlap_data, clusters,
            ranked_df=ranked_df,
            mt5_metrics=mt5_metrics,
            mt5_folder=mt5_folder,
            chart_map=chart_map,
            mc95_data=mc95_data,
            mc_failed=mc_failed,
            mt5_overviews=mt5_overviews,
            keep=keep,
            abandon=abandon,
            strategy_scores=strategy_scores,
            strategy_codes=strategy_codes,
            sqx_metadata=sqx_metadata,
        )
        dashboard_size = os.path.getsize(dashboard_path)
        print(f"  Dashboard saved to: {dashboard_path} ({dashboard_size:,} bytes)")
        if dashboard_size == 0:
            print("  WARNING: Dashboard file is empty - check for errors above.")
        else:
            # Auto-open dashboard in default browser
            import webbrowser
            dashboard_url = 'file:///' + os.path.abspath(dashboard_path).replace('\\', '/')
            print(f"  Opening dashboard in browser...")
            webbrowser.open(dashboard_url)
    except Exception as e:
        import traceback
        print(f"  WARNING: Dashboard generation failed: {e}")
        traceback.print_exc()

    return output_path, keep, abandon, strategy_scores



# ============================================================================
# HTML DASHBOARD GENERATOR
# ============================================================================
def find_equity_chart(strategy_name, folder_path):
    """Find the MT5 equity chart PNG for a strategy."""
    # Try common patterns
    patterns = [
        f"{strategy_name} MT5.png",
        f"{strategy_name}_MT5.png",
        f"{strategy_name}.png",
    ]
    for pat in patterns:
        full = os.path.join(folder_path, pat)
        if os.path.exists(full):
            return pat
    return None


def generate_dashboard(folder_path, strategies, stats, names,
                       corr_daily, corr_weekly, corr_monthly,
                       overlap_data, clusters, ranked_df=None,
                       mt5_metrics=None, mt5_folder=None, chart_map=None,
                       mc95_data=None, mc_failed=None, mt5_overviews=None,
                       keep=None, abandon=None, strategy_scores=None,
                       strategy_codes=None, sqx_metadata=None):
    """Generate the HTML dashboard."""

    def _safe(val, decimals=2, as_int=False):
        """Safely convert a value for JSON — handles NaN, None, numpy types."""
        if val is None:
            return 0
        try:
            v = float(val)
            if v != v or v == float('inf') or v == float('-inf'):
                return 0
            if as_int:
                return int(v)
            return round(v, decimals)
        except (TypeError, ValueError):
            return 0

    n = len(names)

    # Use keep/abandon from caller (already computed using composite score)
    if keep is None or abandon is None:
        # Fallback: compute locally (should not normally happen)
        keep = {}
        abandon = {}
        for cid, members in enumerate(clusters, 1):
            if len(members) == 1:
                keep[members[0]] = cid
            else:
                if strategy_scores:
                    best = max(members, key=lambda s: strategy_scores.get(s, -1))
                else:
                    best = max(members, key=lambda s: stats[s]['total_pnl'])
                keep[best] = cid
                for s in members:
                    if s != best:
                        abandon[s] = (cid, best)

    # Use provided chart_map or empty
    if chart_map is None:
        chart_map = {}
    if mc95_data is None:
        mc95_data = {}

    # Pre-encode equity chart PNGs as base64 data URIs
    # This makes the HTML self-contained (no external PNG dependencies)
    import base64
    chart_b64 = {}
    if chart_map:
        dashboard_dir_tmp = os.path.join(folder_path, 'Dashboard')
        for sname, chart_file in chart_map.items():
            png_path = os.path.join(dashboard_dir_tmp, chart_file)
            if os.path.isfile(png_path):
                try:
                    with open(png_path, 'rb') as cf:
                        b64 = base64.b64encode(cf.read()).decode('ascii')
                        chart_b64[sname] = f'data:image/png;base64,{b64}'
                except Exception:
                    pass

    def _get_mc95(name):
        """Look up MC95 data for a strategy name."""
        matched = match_csv_to_mc(name, set(mc95_data.keys()))
        if matched and matched in mc95_data:
            return mc95_data[matched]
        return {}

    # Build JSON data for JS
    summary_data = []
    for name in names:
        s = stats[name]
        mc = _get_mc95(name)
        summary_data.append({
            'name': name,
            'trades': s['trades'],
            'long': s['long'],
            'short': s['short'],
            'direction': s['direction'],
            'win_rate': round(s['win_rate'] * 100, 1),
            'avg_trade': round(s['avg_trade'], 2),
            'total_pnl': round(s['total_pnl'], 2),
            'max_dd': round(s['max_dd'], 2),
            'dd_days': s['dd_days'],
            'mc95_ret_dd': mc.get('mc95_ret_dd', 0),
            'mc95_ret_dd_tick': None,  # Placeholder for tick-based MC95 Ret/DD
            'chart': chart_b64.get(name, ''),
        })

    ranking_data = []
    if ranked_df is not None and not ranked_df.empty:
        for _, row in ranked_df.iterrows():
            name = row.get('Strategy', '')
            mc = _get_mc95(name)
            ranking_data.append({
                'rank': _safe(row.get('Rank', 0), as_int=True),
                'name': name,
                'score': _safe(row.get('Composite Score', 0), 3),
                'symbol': row.get('Symbol', '') or '',
                'net_profit': _safe(row.get('Total Net Profit', 0)),
                'ret_dd': _safe(row.get('Ret/DD Ratio', 0)),
                'mc95_ret_dd': mc.get('mc95_ret_dd', 0),
                'mc95_ret_dd_tick': None,  # Placeholder for tick-based MC95 Ret/DD (populated by Step9)
                'wl_ratio': _safe(row.get('Win/Loss Ratio', 0)),
                'pf': _safe(row.get('Profit Factor', 0)),
                'sharpe': _safe(row.get('Sharpe Ratio', 0)),
                'recovery': _safe(row.get('Recovery Factor', 0)),
                'lr_corr': _safe(row.get('LR Correlation', 0)),
                'win_rate': _safe(row.get('Win Rate %', 0), 1),
                'trades': _safe(row.get('Total Trades', 0), as_int=True),
                'exp_payoff': _safe(row.get('Expected Payoff', 0)),
                'dd_dollar': _safe(row.get('Balance DD Max $', row.get('Equity DD Max $', 0))),
                'dd_pct': _safe(row.get('Balance DD Rel %', row.get('Equity DD Max %', 0)), 1),
                'lr_stderr': _safe(row.get('LR Standard Error', 0)),
                'chart': chart_b64.get(name, ''),
            })

    portfolio_data = []
    for name in sorted(keep.keys(), key=lambda s: keep[s]):
        cid = keep[name]
        s = stats[name]
        mc = _get_mc95(name)
        cluster_size = len(clusters[cid - 1])
        if cluster_size == 1:
            reason = 'Only in cluster'
        elif strategy_scores and name in strategy_scores:
            reason = f'Best score ({strategy_scores[name]:.3f}) in cluster of {cluster_size}'
        else:
            reason = f'Best P&L in cluster of {cluster_size}'
        portfolio_data.append({
            'name': name, 'decision': 'KEEP', 'cluster': cid,
            'reason': reason,
            'direction': s['direction'],
            'total_pnl': round(s['total_pnl'], 2),
            'avg_trade': round(s['avg_trade'], 2),
            'win_rate': round(s['win_rate'] * 100, 1),
            'mc95_ret_dd': mc.get('mc95_ret_dd', 0),
            'mc95_ret_dd_tick': None,  # Placeholder for tick-based MC95 Ret/DD
            'chart': chart_b64.get(name, ''),
        })
    for name in sorted(abandon.keys(), key=lambda s: abandon[s][0]):
        cid, replaced_by = abandon[name]
        s = stats[name]
        mc = _get_mc95(name)
        if strategy_scores and name in strategy_scores:
            my_score = strategy_scores.get(name, 0)
            their_score = strategy_scores.get(replaced_by, 0)
            reason = f'Correlated with {replaced_by} (score {their_score:.3f} vs {my_score:.3f})'
        else:
            reason = f'Redundant \u2014 correlated with {replaced_by}'
        portfolio_data.append({
            'name': name, 'decision': 'ABANDON', 'cluster': cid,
            'reason': reason,
            'direction': s['direction'],
            'total_pnl': round(s['total_pnl'], 2),
            'avg_trade': round(s['avg_trade'], 2),
            'win_rate': round(s['win_rate'] * 100, 1),
            'mc95_ret_dd': mc.get('mc95_ret_dd', 0),
            'mc95_ret_dd_tick': None,  # Placeholder for tick-based MC95 Ret/DD
            'chart': chart_b64.get(name, ''),
        })

    # Correlation matrices as nested lists
    def corr_to_list(corr_matrix):
        result = []
        for n1 in names:
            row = []
            for n2 in names:
                val = corr_matrix.loc[n1, n2]
                row.append(_safe(val, 3))
            result.append(row)
        return result

    corr_data = {
        'daily': corr_to_list(corr_daily),
        'weekly': corr_to_list(corr_weekly),
        'monthly': corr_to_list(corr_monthly),
    }

    # Overlap pairs
    overlap_list = []
    for (n1, n2), data in sorted(overlap_data.items(), key=lambda x: x[1]['pct'], reverse=True):
        overlap_list.append({
            'pair': f'{n1} vs {n2}',
            'overlaps': data['overlaps'], 'total': data['total'],
            'pct': round(data['pct'] * 100, 1),
            'same': data['same'], 'opp': data['opp'],
        })

    # Clusters
    cluster_data = []
    for i, cluster in enumerate(clusters):
        cluster_data.append({'id': i + 1, 'members': cluster, 'count': len(cluster)})

    # Best pairs (lowest weekly correlation)
    pairs_list = []
    for i, n1 in enumerate(names):
        for j, n2 in enumerate(names):
            if j > i:
                pairs_list.append({
                    'pair': f'{n1} vs {n2}',
                    'weekly': round(float(corr_weekly.loc[n1, n2]), 3),
                    'daily': round(float(corr_daily.loc[n1, n2]), 3),
                })
    pairs_list.sort(key=lambda x: abs(x['weekly']))

    # Drawdown data
    dd_data = []
    for name in names:
        s = stats[name]
        dd_data.append({
            'name': name,
            'max_dd': round(s['max_dd'], 2),
            'dd_start': s['dd_start'].strftime('%Y-%m-%d'),
            'dd_end': s['dd_end'].strftime('%Y-%m-%d'),
            'dd_days': s['dd_days'],
        })

    # Summary stats for cards
    keep_pnl = sum(stats[s]['total_pnl'] for s in keep)
    abandon_pnl = sum(stats[s]['total_pnl'] for s in abandon)

    # Build overviews dict for JS (keyed by strategy name)
    overview_js = {}
    if mt5_overviews:
        for name, ov in mt5_overviews.items():
            overview_js[name] = {
                'symbol': ov.get('symbol', ''),
                'period': ov.get('period', ''),
                'total_profit': round(ov.get('total_profit', 0), 2),
                'gross_profit': round(ov.get('gross_profit', 0), 2),
                'gross_loss': round(ov.get('gross_loss', 0), 2),
                'profit_factor': round(ov.get('profit_factor', 0), 2),
                'expected_payoff': round(ov.get('expected_payoff', 0), 2),
                'recovery_factor': round(ov.get('recovery_factor', 0), 2),
                'sharpe': round(ov.get('sharpe', 0), 2),
                'lr_corr': round(ov.get('lr_corr', 0), 2),
                'lr_stderr': round(ov.get('lr_stderr', 0), 2),
                'total_trades': ov.get('total_trades', 0),
                'avg_win': round(ov.get('avg_win', 0), 2),
                'avg_loss': round(ov.get('avg_loss', 0), 2),
                'largest_win': round(ov.get('largest_win', 0), 2),
                'largest_loss': round(ov.get('largest_loss', 0), 2),
                'bal_dd_max': round(ov.get('bal_dd_max', 0), 2),
                'bal_dd_max_pct': round(ov.get('bal_dd_max_pct', 0), 1),
                'eq_dd_max': round(ov.get('eq_dd_max', 0), 2),
                'eq_dd_max_pct': round(ov.get('eq_dd_max_pct', 0), 1),
                'win_trades': ov.get('win_trades', 0),
                'win_rate': round(ov.get('win_rate', 0), 1),
                'loss_trades': ov.get('loss_trades', 0),
                'short_trades': ov.get('short_trades', 0),
                'short_win_pct': round(ov.get('short_win_pct', 0), 1),
                'long_trades': ov.get('long_trades', 0),
                'long_win_pct': round(ov.get('long_win_pct', 0), 1),
                'max_consec_wins': ov.get('max_consec_wins', 0),
                'max_consec_losses': ov.get('max_consec_losses', 0),
                'avg_consec_wins': ov.get('avg_consec_wins', 0),
                'avg_consec_losses': ov.get('avg_consec_losses', 0),
                'ret_dd': round(ov.get('ret_dd', 0), 2),
                'wl_ratio': round(ov.get('wl_ratio', 0), 2),
                'payout_ratio': round(ov.get('payout_ratio', 0), 2),
                'dd_dollar': round(ov.get('dd_dollar', 0), 2),
                'dd_pct': round(ov.get('dd_pct', 0), 1),
                'monthly_avg_profit': round(ov.get('monthly_avg_profit', 0), 2),
                'cagr': ov.get('cagr', 0),
                'stagnation_days': ov.get('stagnation_days', 0),
                'stagnation_pct': ov.get('stagnation_pct', 0),
                'yearly_avg_profit': round(ov.get('yearly_avg_profit', 0), 2),
                'yearly_avg_return_pct': round(ov.get('yearly_avg_return_pct', 0), 2),
                'daily_avg_profit': round(ov.get('daily_avg_profit', 0), 2),
                'annual_dd_ratio': ov.get('annual_dd_ratio', 0),
                'monthly_pnl': ov.get('monthly_pnl', {}),
            }

    # Strategy pseudo code data
    codes_js = {}
    if strategy_codes:
        for name, code in strategy_codes.items():
            codes_js[name] = {
                'direction': code.get('direction', ''),
                'style': code.get('style', ''),
                'entry_type': code.get('entry_type', ''),
                'indicators': code.get('indicators', []),
                'entry_refs': code.get('entry_refs', []),
                'trailing_stop': code.get('trailing_stop', False),
                'ts_activation': code.get('ts_activation', False),
                'move_sl_be': code.get('move_sl_be', False),
                'profit_target': code.get('profit_target', False),
                'exit_after_bars': code.get('exit_after_bars'),
                'has_exit_signals': code.get('has_exit_signals', False),
                'time_filter': code.get('time_filter', ''),
                'order_valid_bars': code.get('order_valid_bars'),
                'exit_summary': code.get('exit_summary', ''),
                'raw_content': code.get('raw_content', ''),
            }

    # SQX metadata (complexity etc.)
    sqx_meta_js = {}
    if sqx_metadata:
        for name, meta in sqx_metadata.items():
            sqx_meta_js[name] = {
                'complexity': meta.get('complexity'),
            }

    js_data = {
        'names': names,
        'summary': summary_data,
        'ranking': ranking_data,
        'portfolio': portfolio_data,
        'mc_failed': mc_failed or [],
        'overviews': overview_js,
        'strategy_codes': codes_js,
        'sqx_metadata': sqx_meta_js,
        'correlations': corr_data,
        'overlap': overlap_list,
        'clusters': cluster_data,
        'best_pairs': pairs_list,
        'drawdowns': dd_data,
        'cards': {
            'total_strategies': n,
            'clusters': len(clusters),
            'keep': len(keep),
            'abandon': len(abandon),
            'keep_pnl': round(keep_pnl, 0),
            'abandon_pnl': round(abandon_pnl, 0),
            'ranked': len(ranking_data),
            'mc_failed': len(mc_failed or []),
        },
    }

    # Create Dashboard folder
    dashboard_dir = os.path.join(folder_path, 'Dashboard')
    os.makedirs(dashboard_dir, exist_ok=True)

    # Build HTML first, then write (avoids 0-byte file on error)
    html_content = build_html(js_data)
    html_path = os.path.join(dashboard_dir, 'index.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    # Save strategy data as JSON for later updates (e.g., tick-based MC95)
    # This file can be read and updated by Step9, then used to regenerate the dashboard
    strategies_json_path = os.path.join(dashboard_dir, 'strategies_data.json')
    strategies_export = {
        'ranking': js_data.get('ranking', []),
        'portfolio': js_data.get('portfolio', []),
        'strategies': js_data.get('strategies', []),
        'mc_failed': js_data.get('mc_failed', []),
        'clusters': js_data.get('clusters', []),
        'strategy_names': js_data.get('strategy_names', []),
        'mc95_threshold': js_data.get('mc95_threshold', 2.5),
        'generated_at': js_data.get('generated_at', ''),
    }
    with open(strategies_json_path, 'w', encoding='utf-8') as f:
        json.dump(strategies_export, f, cls=_SafeJSONEncoder, indent=2)
    print(f"  Saved strategies data to: {strategies_json_path}")

    # Clean up equity chart PNGs — they're now embedded as base64 in the HTML
    png_count = 0
    for fname in os.listdir(dashboard_dir):
        if fname.endswith('_equity.png'):
            try:
                os.remove(os.path.join(dashboard_dir, fname))
                png_count += 1
            except Exception:
                pass
    if png_count:
        print(f"  Cleaned up {png_count} equity chart PNG(s) (embedded in HTML)")

    return html_path


class _SafeJSONEncoder(json.JSONEncoder):
    """JSON encoder that handles NaN, inf, numpy types."""
    def default(self, obj):
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            v = float(obj)
            if np.isnan(v) or np.isinf(v):
                return None
            return v
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return super().default(obj)

    def encode(self, obj):
        return super().encode(self._sanitize(obj))

    def _sanitize(self, obj):
        if isinstance(obj, float):
            if obj != obj or obj == float('inf') or obj == float('-inf'):
                return None
            return obj
        if isinstance(obj, dict):
            return {k: self._sanitize(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [self._sanitize(v) for v in obj]
        return obj


def build_html(data):
    """Build the complete HTML dashboard string."""
    data_json = json.dumps(data, cls=_SafeJSONEncoder)

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Strategy Correlation Analysis</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;500;600;700&family=Outfit:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
{CSS_BLOCK}
</style>
</head>
<body>

<div id="app">
  <header class="top-bar">
    <div class="logo">
      <svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 3v18h18"/><path d="M7 16l4-8 4 5 4-9"/></svg>
      <span>Strategy Analysis</span>
    </div>
    <nav class="tabs" id="mainTabs">
      <button class="tab active" data-tab="ranking">Rankings</button>
      <button class="tab" data-tab="mc-failed">MC Failed</button>
      <button class="tab" data-tab="portfolio">Portfolio</button>
      <button class="tab" data-tab="correlation">Correlation</button>
      <button class="tab" data-tab="overlap">Overlap</button>
      <button class="tab" data-tab="clusters">Clusters</button>
      <button class="tab" data-tab="drawdowns">Drawdowns</button>
      <button class="tab" data-tab="methodology">Methodology</button>
    </nav>
  </header>

  <!-- Summary Cards -->
  <div class="cards" id="summaryCards"></div>

  <!-- Main content area with optional side panel -->
  <div class="main-layout">
    <div class="content-area">
      <div class="panel" id="panel-ranking"></div>
      <div class="panel hidden" id="panel-mc-failed"></div>
      <div class="panel hidden" id="panel-portfolio"></div>
      <div class="panel hidden" id="panel-correlation"></div>
      <div class="panel hidden" id="panel-overlap"></div>
      <div class="panel hidden" id="panel-clusters"></div>
      <div class="panel hidden" id="panel-drawdowns"></div>
      <div class="panel hidden" id="panel-methodology"></div>
    </div>
    <div class="side-panel hidden" id="sidePanel">
      <div class="side-panel-header">
        <h3 id="sidePanelTitle">Equity Curve</h3>
        <button class="close-btn" id="closeSidePanel">&times;</button>
      </div>
      <div class="side-panel-body" id="sidePanelBody"></div>
    </div>
  </div>
</div>

<!-- Chart lightbox modal -->
<div class="chart-lightbox" id="chartLightbox">
  <div class="lb-title" id="lbTitle"></div>
  <div class="lb-close" id="lbClose">&times;</div>
  <img id="lbImg" src="" alt="">
</div>

<script>
const DATA = {data_json};
{JS_BLOCK}
</script>
</body>
</html>'''


# ============================================================================
# CSS
# ============================================================================
CSS_BLOCK = r"""
:root {
  --bg-primary: #0c0e14;
  --bg-secondary: #12151e;
  --bg-card: #181c28;
  --bg-hover: #1e2333;
  --bg-active: #252a3a;
  --border: #2a2f42;
  --border-light: #353b52;
  --text-primary: #e8eaf0;
  --text-secondary: #8b90a5;
  --text-muted: #5c6178;
  --accent: #6c9eff;
  --accent-dim: rgba(108,158,255,0.12);
  --green: #4ade80;
  --green-dim: rgba(74,222,128,0.12);
  --red: #f87171;
  --red-dim: rgba(248,113,113,0.12);
  --amber: #fbbf24;
  --amber-dim: rgba(251,191,36,0.12);
  --purple: #a78bfa;
  --tier-1: rgba(74,222,128,0.10);
  --tier-1-border: rgba(74,222,128,0.25);
  --tier-2: rgba(108,158,255,0.10);
  --tier-2-border: rgba(108,158,255,0.25);
  --tier-3: rgba(251,191,36,0.08);
  --tier-3-border: rgba(251,191,36,0.20);
  --font-body: 'Outfit', sans-serif;
  --font-mono: 'JetBrains Mono', monospace;
  --radius: 8px;
  --radius-lg: 12px;
}

* { margin:0; padding:0; box-sizing:border-box; }

html { font-size: 14px; }

body {
  font-family: var(--font-body);
  background: var(--bg-primary);
  color: var(--text-primary);
  line-height: 1.5;
  min-height: 100vh;
}

#app { max-width: 1600px; margin: 0 auto; padding: 0 24px 48px; }

/* Top bar */
.top-bar {
  display: flex; align-items: center; gap: 32px;
  padding: 16px 0; border-bottom: 1px solid var(--border);
  margin-bottom: 20px; position: sticky; top: 0;
  background: var(--bg-primary); z-index: 100;
}
.logo {
  display: flex; align-items: center; gap: 8px;
  font-weight: 700; font-size: 1.1rem; color: var(--accent);
  white-space: nowrap;
}
.tabs { display: flex; gap: 2px; }
.tab {
  background: none; border: none; color: var(--text-secondary);
  font-family: var(--font-body); font-size: 0.85rem; font-weight: 500;
  padding: 8px 16px; border-radius: var(--radius); cursor: pointer;
  transition: all 0.15s;
}
.tab:hover { color: var(--text-primary); background: var(--bg-hover); }
.tab.active { color: var(--accent); background: var(--accent-dim); }

/* Cards */
.cards {
  display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
  gap: 12px; margin-bottom: 20px;
}
.card {
  background: var(--bg-card); border: 1px solid var(--border);
  border-radius: var(--radius-lg); padding: 16px 18px;
}
.card-label { font-size: 0.75rem; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; font-weight: 600; }
.card-value { font-family: var(--font-mono); font-size: 1.5rem; font-weight: 700; margin-top: 4px; }
.card-value.green { color: var(--green); }
.card-value.red { color: var(--red); }
.card-value.accent { color: var(--accent); }
.card-value.amber { color: var(--amber); }

/* Layout */
.main-layout { display: flex; gap: 0; }
.content-area { flex: 1; min-width: 0; }
.panel { animation: fadeIn 0.2s ease; }
.hidden { display: none !important; }

@keyframes fadeIn { from { opacity: 0; transform: translateY(6px); } to { opacity: 1; transform: translateY(0); } }

/* Side Panel */
.side-panel {
  width: 520px; min-width: 520px;
  background: var(--bg-secondary); border-left: 1px solid var(--border);
  margin-left: 16px; border-radius: var(--radius-lg);
  position: sticky; top: 76px; max-height: calc(100vh - 96px);
  overflow-y: auto; animation: slideIn 0.2s ease;
}
@keyframes slideIn { from { opacity: 0; transform: translateX(12px); } to { opacity: 1; transform: translateX(0); } }
.side-panel-header {
  display: flex; justify-content: space-between; align-items: center;
  padding: 16px 20px; border-bottom: 1px solid var(--border);
  position: sticky; top: 0; background: var(--bg-secondary); z-index: 2;
}
.side-panel-header h3 { font-size: 0.9rem; font-weight: 600; color: var(--text-primary); }
.close-btn {
  background: none; border: none; color: var(--text-muted); font-size: 1.4rem;
  cursor: pointer; padding: 2px 6px; border-radius: 4px; line-height: 1;
}
.close-btn:hover { color: var(--text-primary); background: var(--bg-hover); }
.side-panel-body { padding: 16px 20px; }
.side-panel-body img {
  width: 100%; border-radius: var(--radius); border: 1px solid var(--border);
  cursor: pointer; transition: opacity 0.15s;
}
.side-panel-body img:hover { opacity: 0.85; }
.side-panel-body .sp-stats { margin-top: 14px; }
.side-panel-body .sp-row {
  display: flex; justify-content: space-between; padding: 5px 0;
  border-bottom: 1px solid var(--border); font-size: 0.82rem;
}
.side-panel-body .sp-row:last-child { border-bottom: none; }
.sp-label { color: var(--text-secondary); }
.sp-val { font-family: var(--font-mono); font-weight: 500; }

/* Lightbox modal for enlarged charts */
.chart-lightbox {
  display: none; position: fixed; inset: 0; z-index: 9999;
  background: rgba(0,0,0,0.85); backdrop-filter: blur(4px);
  justify-content: center; align-items: center; cursor: zoom-out;
}
.chart-lightbox.active { display: flex; }
.chart-lightbox img {
  max-width: 92vw; max-height: 88vh; border-radius: 8px;
  border: 1px solid var(--border); box-shadow: 0 20px 60px rgba(0,0,0,0.5);
}
.chart-lightbox .lb-title {
  position: absolute; top: 20px; left: 50%; transform: translateX(-50%);
  color: #e8eaf0; font-size: 1rem; font-weight: 600; padding: 6px 16px;
  background: rgba(18,21,30,0.8); border-radius: 6px;
}
.chart-lightbox .lb-close {
  position: absolute; top: 16px; right: 20px; color: #8b90a5; font-size: 1.6rem;
  cursor: pointer; width: 36px; height: 36px; display: flex; align-items: center;
  justify-content: center; border-radius: 50%; background: rgba(18,21,30,0.6);
  transition: all 0.15s;
}
.chart-lightbox .lb-close:hover { color: #fff; background: rgba(255,255,255,0.1); }

/* Methodology panel */
.meth-section { margin-bottom: 20px; }
.meth-heading {
  font-size: 0.95rem; font-weight: 600; color: var(--accent);
  margin-bottom: 8px; padding-bottom: 4px; border-bottom: 1px solid var(--border);
}
.meth-section p {
  color: var(--text-secondary); font-size: 0.85rem; line-height: 1.6; margin: 0;
}
.meth-table {
  width: 100%; border-collapse: collapse; font-size: 0.82rem; margin-top: 6px;
}
.meth-table th {
  text-align: left; padding: 8px 10px; background: var(--bg-hover);
  color: var(--text-primary); font-weight: 600; border-bottom: 2px solid var(--border);
}
.meth-table td {
  padding: 6px 10px; border-bottom: 1px solid var(--border); color: var(--text-secondary);
}
.meth-table tr:hover td { background: rgba(108,158,255,0.04); }
.meth-colours { display: flex; align-items: center; gap: 4px; font-size: 0.85rem; color: var(--text-secondary); margin-top: 6px; }
.meth-swatch {
  display: inline-block; width: 20px; height: 14px; border-radius: 3px; border: 1px solid var(--border);
}
.tier-1-bg { background: rgba(0, 184, 148, 0.15); }
.tier-2-bg { background: rgba(108, 158, 255, 0.12); }
.tier-3-bg { background: rgba(251, 191, 36, 0.12); }
.meth-notes {
  margin: 6px 0 0 0; padding-left: 20px; font-size: 0.85rem;
  color: var(--text-secondary); line-height: 1.8;
}
.meth-notes code {
  background: var(--bg-hover); padding: 1px 5px; border-radius: 3px;
  font-family: var(--font-mono); font-size: 0.8rem; color: var(--text-primary);
}

/* Tables */
.table-wrap {
  overflow-x: auto; border: 1px solid var(--border);
  border-radius: var(--radius-lg); background: var(--bg-card);
}
table {
  width: 100%; border-collapse: collapse; font-size: 0.82rem;
}
th {
  text-align: left; padding: 10px 12px; font-weight: 600;
  color: var(--text-muted); text-transform: uppercase; font-size: 0.7rem;
  letter-spacing: 0.04em; border-bottom: 1px solid var(--border);
  background: var(--bg-secondary); position: sticky; top: 0; z-index: 1;
  cursor: pointer; user-select: none; white-space: nowrap;
}
th:hover { color: var(--text-secondary); }
th.sorted-asc::after { content: ' ▲'; color: var(--accent); }
th.sorted-desc::after { content: ' ▼'; color: var(--accent); }
th.no-sort { cursor: default; }
td {
  padding: 8px 12px; border-bottom: 1px solid var(--border);
  white-space: nowrap; font-family: var(--font-mono); font-size: 0.8rem;
}
td.text-col { font-family: var(--font-body); }
tr:hover td { background: var(--bg-hover); }
tr:last-child td { border-bottom: none; }
tr.tier-1 td { background: var(--tier-1); }
tr.tier-1:hover td { background: var(--tier-1-border); }
tr.tier-2 td { background: var(--tier-2); }
tr.tier-2:hover td { background: var(--tier-2-border); }
tr.tier-3 td { background: var(--tier-3); }
tr.tier-3:hover td { background: var(--tier-3-border); }

/* Abandoned rows - dimmed */
tr.abandon-row td { opacity: 0.45; }
tr.abandon-row:hover td { opacity: 0.75; }
.abandon-name { color: var(--text-muted); }

/* Chart button */
.analyze-btn {
  background: var(--accent-dim); border: 1px solid rgba(108,158,255,0.2);
  color: var(--accent); border-radius: 4px; padding: 3px 8px;
  cursor: pointer; font-size: 0.72rem; font-family: var(--font-body);
  font-weight: 500; transition: all 0.15s; white-space: nowrap;
}
.analyze-btn:hover { background: rgba(108,158,255,0.2); border-color: var(--accent); }
.analyze-btn.active { background: var(--accent); color: var(--bg-primary); }
.analyze-btn.disabled { opacity: 0.3; cursor: not-allowed; }
.analyze-btn.ov-btn { color: #fbbf24; background: rgba(251,191,36,0.08); border-color: rgba(251,191,36,0.2); }
.analyze-btn.ov-btn:hover { background: rgba(251,191,36,0.18); border-color: #fbbf24; }
.analyze-btn.ov-btn.active { background: #fbbf24; color: var(--bg-primary); }
.analyze-btn.export-btn { color: #34d399; background: rgba(52,211,153,0.08); border-color: rgba(52,211,153,0.2); }
.analyze-btn.export-btn:hover { background: rgba(52,211,153,0.18); border-color: #34d399; }
.analyze-btn.code-btn { color: #a78bfa; background: rgba(167,139,250,0.08); border-color: rgba(167,139,250,0.2); }
.analyze-btn.code-btn:hover { background: rgba(167,139,250,0.18); border-color: #a78bfa; }
.analyze-btn.code-btn.active { background: #a78bfa; color: var(--bg-primary); }
.btn-group { display: flex; gap: 4px; }
/* Keep old class for backwards compat */
.chart-btn { /* alias handled by analyze-btn */ }

/* Overview panel */
.ov-grid { margin-bottom: 16px; }
.ov-hero {
  text-align: center; padding: 12px; margin-bottom: 10px;
  background: var(--bg-hover); border-radius: var(--radius); border: 1px solid var(--border);
}
.ov-hero-label { font-size: 0.7rem; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; }
.ov-hero-val { font-size: 1.5rem; font-weight: 700; font-family: var(--font-mono); }
.ov-hero-val.pos { color: var(--green); }
.ov-hero-val.neg { color: var(--red); }
.ov-kpi-row { display: flex; gap: 6px; margin-bottom: 6px; }
.ov-kpi {
  flex: 1; text-align: center; padding: 8px 4px;
  background: var(--bg-hover); border-radius: var(--radius); border: 1px solid var(--border);
}
.ov-kpi-label { font-size: 0.62rem; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.03em; }
.ov-kpi-val { font-size: 0.85rem; font-weight: 600; font-family: var(--font-mono); color: var(--text-primary); }
.ov-section { margin-top: 14px; }
.ov-section-title {
  font-size: 0.8rem; font-weight: 600; color: var(--accent);
  margin-bottom: 6px; padding-bottom: 4px; border-bottom: 1px solid var(--border);
}
.sp-val.pos, .sp-val .pos { color: var(--green); }
.sp-val.neg, .sp-val .neg { color: var(--red); }

/* Strategy code panel */
.code-tag { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 0.72rem; font-weight: 600; margin-right: 4px; margin-bottom: 3px; }
.code-tag.style { background: rgba(167,139,250,0.15); color: #a78bfa; }
.code-tag.direction { background: rgba(59,130,246,0.15); color: #60a5fa; }
.code-tag.entry-type { background: rgba(251,191,36,0.15); color: #fbbf24; }
.code-tag.yes { background: rgba(34,197,94,0.12); color: var(--green); }
.code-tag.no { background: rgba(239,68,68,0.08); color: var(--text-muted); }
.code-indicators { display: flex; flex-wrap: wrap; gap: 3px; margin-top: 4px; }
.code-ind-tag { display: inline-block; padding: 2px 7px; border-radius: 3px; font-size: 0.68rem;
  background: rgba(99,102,241,0.1); color: #818cf8; border: 1px solid rgba(99,102,241,0.2); }
.code-ref-tag { display: inline-block; padding: 2px 7px; border-radius: 3px; font-size: 0.68rem;
  background: rgba(251,191,36,0.1); color: #fbbf24; border: 1px solid rgba(251,191,36,0.2); }

/* Monthly P&L grid */
.monthly-grid-wrap { overflow-x: auto; }
.monthly-grid {
  width: 100%; border-collapse: collapse; font-size: 0.68rem;
  font-family: var(--font-mono);
}
.monthly-grid th {
  padding: 3px 2px; font-weight: 600; color: var(--text-muted);
  border-bottom: 1px solid var(--border); text-align: center; font-size: 0.65rem;
}
.monthly-grid td {
  padding: 2px 2px; text-align: right; border-bottom: 1px solid rgba(42,47,66,0.5);
  white-space: nowrap;
}
.monthly-grid td.pos { color: var(--green); }
.monthly-grid td.neg { color: var(--red); }
.monthly-grid td.yr { text-align: center; font-weight: 600; color: var(--text-primary); }
.monthly-grid td.ytd { font-weight: 600; border-left: 1px solid var(--border); }

/* Portfolio tags */
.tag {
  display: inline-block; padding: 2px 10px; border-radius: 20px;
  font-size: 0.72rem; font-weight: 600; letter-spacing: 0.02em;
}
.tag-keep { background: var(--green-dim); color: var(--green); border: 1px solid rgba(74,222,128,0.3); }
.tag-abandon { background: var(--red-dim); color: var(--red); border: 1px solid rgba(248,113,113,0.3); }
.tag-long { color: #60a5fa; }
.tag-short { color: #f87171; }
.tag-both { color: var(--purple); }

/* Correlation heatmap */
.heatmap-controls { display: flex; gap: 8px; margin-bottom: 14px; }
.hm-btn {
  background: var(--bg-card); border: 1px solid var(--border); color: var(--text-secondary);
  padding: 6px 14px; border-radius: var(--radius); font-family: var(--font-body);
  font-size: 0.8rem; cursor: pointer; transition: all 0.15s;
}
.hm-btn:hover { border-color: var(--border-light); color: var(--text-primary); }
.hm-btn.active { background: var(--accent-dim); border-color: rgba(108,158,255,0.3); color: var(--accent); }

.heatmap-grid {
  display: grid; gap: 2px; overflow-x: auto;
  padding: 4px; background: var(--bg-card); border-radius: var(--radius-lg);
  border: 1px solid var(--border);
}
.hm-cell {
  display: flex; align-items: center; justify-content: center;
  font-family: var(--font-mono); font-size: 0.7rem; font-weight: 500;
  border-radius: 3px; min-width: 52px; height: 32px;
  cursor: default; transition: transform 0.1s;
}
.hm-cell:hover { transform: scale(1.08); z-index: 1; }
.hm-header {
  font-family: var(--font-body); font-size: 0.65rem; font-weight: 600;
  color: var(--text-muted); writing-mode: vertical-lr; text-align: center;
  transform: rotate(180deg); min-height: 80px; padding: 4px;
  display: flex; align-items: center; justify-content: center;
}
.hm-row-header {
  font-family: var(--font-body); font-size: 0.68rem; font-weight: 500;
  color: var(--text-secondary); text-align: right; padding-right: 8px;
  display: flex; align-items: center; justify-content: flex-end;
  overflow: hidden; text-overflow: ellipsis;
}

/* Colour key */
.colour-key {
  display: flex; gap: 12px; margin-top: 12px; font-size: 0.75rem;
  color: var(--text-secondary);
}
.colour-key span {
  display: inline-flex; align-items: center; gap: 5px;
}
.ck-swatch {
  width: 14px; height: 14px; border-radius: 3px;
  display: inline-block;
}

/* Section headers */
.section-header {
  font-size: 1.1rem; font-weight: 700; color: var(--text-primary);
  margin-bottom: 14px; display: flex; align-items: center; gap: 8px;
}
.section-sub {
  font-size: 0.82rem; color: var(--text-muted); margin-bottom: 16px;
  font-weight: 400;
}

/* Cluster cards */
.cluster-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 12px; }
.cluster-card {
  background: var(--bg-card); border: 1px solid var(--border);
  border-radius: var(--radius-lg); padding: 16px;
}
.cluster-card h4 {
  font-size: 0.85rem; font-weight: 600; color: var(--accent);
  margin-bottom: 8px;
}
.cluster-member {
  font-size: 0.8rem; color: var(--text-secondary); padding: 3px 0;
  font-family: var(--font-mono);
}
.cluster-member.is-best { color: var(--green); font-weight: 600; }

/* Scrollbar */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: var(--border-light); }
"""

# ============================================================================
# JavaScript
# ============================================================================
JS_BLOCK = r"""
(function() {
  const $ = s => document.querySelector(s);
  const $$ = s => document.querySelectorAll(s);
  let activeSideStrategy = null;

  // === Tab switching ===
  $$('.tab').forEach(btn => {
    btn.addEventListener('click', () => {
      $$('.tab').forEach(b => b.classList.remove('active'));
      btn.classList.add('active');
      $$('.panel').forEach(p => p.classList.add('hidden'));
      $(`#panel-${btn.dataset.tab}`).classList.remove('hidden');
    });
  });

  // === Summary Cards ===
  const c = DATA.cards;
  $('#summaryCards').innerHTML = `
    <div class="card"><div class="card-label">Strategies</div><div class="card-value accent">${c.total_strategies}</div></div>
    <div class="card"><div class="card-label">Clusters</div><div class="card-value">${c.clusters}</div></div>
    <div class="card"><div class="card-label">Keep</div><div class="card-value green">${c.keep}</div></div>
    <div class="card"><div class="card-label">Abandon</div><div class="card-value red">${c.abandon}</div></div>
    <div class="card"><div class="card-label">Keep P&L</div><div class="card-value green">$${c.keep_pnl.toLocaleString()}</div></div>
    ${c.ranked ? `<div class="card"><div class="card-label">MT5 Ranked</div><div class="card-value amber">${c.ranked}</div></div>` : ''}
    ${c.mc_failed ? `<div class="card"><div class="card-label">MC Failed</div><div class="card-value red">${c.mc_failed}</div></div>` : ''}
  `;

  // === Side Panel ===
  function openSidePanel(name, chartSrc, extraStats) {
    const sp = $('#sidePanel');
    const body = $('#sidePanelBody');
    activeSideStrategy = name;
    $('#sidePanelTitle').textContent = name;

    let html = '';
    if (chartSrc) {
      html += `<img src="${chartSrc}" alt="Equity curve for ${name}" loading="lazy" onclick="openLightbox('${chartSrc}', '${name.replace(/'/g, "\\'")}')" title="Click to enlarge">`;
    } else {
      html += `<div style="padding:40px;text-align:center;color:var(--text-muted);border:1px dashed var(--border);border-radius:var(--radius);">No equity chart available</div>`;
    }
    if (extraStats) {
      html += '<div class="sp-stats">';
      for (const [label, val] of extraStats) {
        html += `<div class="sp-row"><span class="sp-label">${label}</span><span class="sp-val">${val}</span></div>`;
      }
      html += '</div>';
    }
    body.innerHTML = html;
    sp.classList.remove('hidden');
    // highlight active chart buttons
    $$('.chart-btn').forEach(b => b.classList.toggle('active', b.dataset.name === name));
  }

  function closeSidePanel() {
    $('#sidePanel').classList.add('hidden');
    activeSideStrategy = null;
    $$('.chart-btn').forEach(b => b.classList.remove('active'));
  }
  $('#closeSidePanel').addEventListener('click', closeSidePanel);

  // === Overview Panel (QA4-style summary) ===
  function openOverviewPanel(name) {
    const sp = $('#sidePanel');
    const body = $('#sidePanelBody');
    const ov = DATA.overviews[name];
    if (!ov) return;

    $('#sidePanelTitle').textContent = name;

    const fmt = v => typeof v === 'number' ? '$' + v.toLocaleString(undefined, {minimumFractionDigits:2}) : v;
    const n = v => typeof v === 'number' ? v.toLocaleString(undefined, {minimumFractionDigits:2}) : v;
    const pct = v => v + '%';

    // Look up MC95 Ret/DD from ranking data
    const rk = DATA.ranking.find(x => x.name === name);
    const mc95val = rk && rk.mc95_ret_dd ? rk.mc95_ret_dd : null;

    let html = `<div class="ov-grid">
      <div class="ov-hero">
        <div class="ov-hero-label">TOTAL PROFIT</div>
        <div class="ov-hero-val ${ov.total_profit >= 0 ? 'pos' : 'neg'}">${fmt(ov.total_profit)}</div>
      </div>
      <div class="ov-kpi-row">
        <div class="ov-kpi"><div class="ov-kpi-label"># Trades</div><div class="ov-kpi-val">${ov.total_trades}</div></div>
        <div class="ov-kpi"><div class="ov-kpi-label">Profit Factor</div><div class="ov-kpi-val">${ov.profit_factor}</div></div>
        <div class="ov-kpi"><div class="ov-kpi-label">Ret / DD</div><div class="ov-kpi-val">${ov.ret_dd}</div></div>
        <div class="ov-kpi"><div class="ov-kpi-label">MC95 Ret/DD</div><div class="ov-kpi-val">${mc95val || '—'}</div></div>
        <div class="ov-kpi"><div class="ov-kpi-label">Win %</div><div class="ov-kpi-val">${pct(ov.win_rate)}</div></div>
      </div>
      <div class="ov-kpi-row">
        <div class="ov-kpi"><div class="ov-kpi-label">Drawdown</div><div class="ov-kpi-val">${fmt(ov.dd_dollar)}</div></div>
        <div class="ov-kpi"><div class="ov-kpi-label">DD %</div><div class="ov-kpi-val">${pct(ov.dd_pct)}</div></div>
        <div class="ov-kpi"><div class="ov-kpi-label">CAGR</div><div class="ov-kpi-val">${pct(ov.cagr)}</div></div>
        <div class="ov-kpi"><div class="ov-kpi-label">Recovery</div><div class="ov-kpi-val">${ov.recovery_factor}</div></div>
        <div class="ov-kpi"><div class="ov-kpi-label">Sharpe</div><div class="ov-kpi-val">${ov.sharpe}</div></div>
      </div>
    </div>`;

    // Stats section
    html += `<div class="ov-section"><div class="ov-section-title">Strategy</div>
      <div class="sp-stats">
        <div class="sp-row"><span class="sp-label">Wins / Losses Ratio</span><span class="sp-val">${ov.wl_ratio}</span></div>
        <div class="sp-row"><span class="sp-label">Payout Ratio (Avg Win/Loss)</span><span class="sp-val">${ov.payout_ratio}</span></div>
        <div class="sp-row"><span class="sp-label">Average Trade</span><span class="sp-val">${fmt(ov.expected_payoff)}</span></div>
        <div class="sp-row"><span class="sp-label">Yearly Avg Profit</span><span class="sp-val">${fmt(ov.yearly_avg_profit)} (${ov.yearly_avg_return_pct}%)</span></div>
        <div class="sp-row"><span class="sp-label">Monthly Avg Profit</span><span class="sp-val">${fmt(ov.monthly_avg_profit)}</span></div>
        <div class="sp-row"><span class="sp-label">Daily Avg Profit</span><span class="sp-val">${fmt(ov.daily_avg_profit)}</span></div>
        <div class="sp-row"><span class="sp-label">Annual% / MaxDD%</span><span class="sp-val">${ov.annual_dd_ratio}</span></div>
        <div class="sp-row"><span class="sp-label">Stagnation</span><span class="sp-val">${ov.stagnation_days} days (${ov.stagnation_pct}%)</span></div>
        <div class="sp-row"><span class="sp-label">LR Standard Error</span><span class="sp-val">${n(ov.lr_stderr)}</span></div>
      </div>
    </div>`;

    // Trades section
    html += `<div class="ov-section"><div class="ov-section-title">Trades</div>
      <div class="sp-stats">
        <div class="sp-row"><span class="sp-label">Gross Profit / Loss</span><span class="sp-val"><span class="pos">${fmt(ov.gross_profit)}</span> / <span class="neg">${fmt(ov.gross_loss)}</span></span></div>
        <div class="sp-row"><span class="sp-label">Average Win / Loss</span><span class="sp-val"><span class="pos">${fmt(ov.avg_win)}</span> / <span class="neg">${fmt(ov.avg_loss)}</span></span></div>
        <div class="sp-row"><span class="sp-label">Largest Win / Loss</span><span class="sp-val"><span class="pos">${fmt(ov.largest_win)}</span> / <span class="neg">${fmt(ov.largest_loss)}</span></span></div>
        <div class="sp-row"><span class="sp-label">Long / Short (Win%)</span><span class="sp-val">${ov.long_trades} (${pct(ov.long_win_pct)}) / ${ov.short_trades} (${pct(ov.short_win_pct)})</span></div>
        <div class="sp-row"><span class="sp-label">Max Consec Wins / Losses</span><span class="sp-val"><span class="pos">${ov.max_consec_wins}</span> / <span class="neg">${ov.max_consec_losses}</span></span></div>
        <div class="sp-row"><span class="sp-label">Avg Consec Wins / Losses</span><span class="sp-val"><span class="pos">${typeof ov.avg_consec_wins === 'number' ? ov.avg_consec_wins.toFixed(2) : ov.avg_consec_wins}</span> / <span class="neg">${typeof ov.avg_consec_losses === 'number' ? ov.avg_consec_losses.toFixed(2) : ov.avg_consec_losses}</span></span></div>
      </div>
    </div>`;

    // Monthly P&L grid
    const mp = ov.monthly_pnl;
    const years = Object.keys(mp).sort().reverse();
    if (years.length) {
      const months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
      html += `<div class="ov-section"><div class="ov-section-title">Monthly Performance ($)</div>
        <div class="monthly-grid-wrap"><table class="monthly-grid">
          <thead><tr><th>Year</th>${months.map(m => '<th>' + m + '</th>').join('')}<th>YTD</th></tr></thead>
          <tbody>`;
      for (const yr of years) {
        const row = mp[yr];
        html += '<tr><td class="yr">' + yr + '</td>';
        for (let m = 1; m <= 12; m++) {
          const v = row[m] || 0;
          const cls = v > 0 ? 'pos' : v < 0 ? 'neg' : '';
          html += '<td class="' + cls + '">' + (v !== 0 ? v.toFixed(0) : '') + '</td>';
        }
        const ytd = row.ytd || 0;
        const ytdCls = ytd > 0 ? 'pos' : ytd < 0 ? 'neg' : '';
        html += '<td class="ytd ' + ytdCls + '">' + ytd.toFixed(0) + '</td></tr>';
      }
      html += '</tbody></table></div></div>';
    }

    body.innerHTML = html;
    sp.classList.remove('hidden');
    $$('.analyze-btn').forEach(b => b.classList.toggle('active',
      b.dataset.name === name && b.dataset.action === 'overview'));
  }

  // === Strategy Code Panel ===
  function openCodePanel(name) {
    const code = DATA.strategy_codes[name];
    if (!code) { alert('No pseudo code available for ' + name); return; }
    const sp = $('#sidePanel');
    const body = $('#sidePanelBody');
    $('#sidePanelTitle').textContent = name;
    activeSideStrategy = name + '_code';

    let html = '<div class="ov-grid" style="padding:12px 0">';
    html += '<div style="margin-bottom:10px">';
    html += `<span class="code-tag direction">${code.direction}</span>`;
    html += `<span class="code-tag style">${code.style}</span>`;
    html += `<span class="code-tag entry-type">${code.entry_type} Order</span>`;
    html += '</div>';

    html += '<div class="ov-section"><div class="ov-section-title">Risk Management</div><div class="sp-stats">';
    const yesNo = (v, label) => `<div class="sp-row"><span class="sp-label">${label}</span><span class="sp-val">${v ? '<span class="code-tag yes">Yes</span>' : '<span class="code-tag no">No</span>'}</span></div>`;
    html += yesNo(true, 'Stop Loss');
    html += yesNo(code.profit_target, 'Profit Target');
    let tsLabel = 'Trailing Stop';
    if (code.trailing_stop && code.ts_activation) tsLabel = 'Trailing Stop (with activation)';
    html += yesNo(code.trailing_stop, tsLabel);
    html += yesNo(code.move_sl_be, 'Move SL to Breakeven');
    if (code.exit_after_bars) {
      html += `<div class="sp-row"><span class="sp-label">Exit After N Bars</span><span class="sp-val"><span class="code-tag yes">${code.exit_after_bars} bars</span></span></div>`;
    } else {
      html += yesNo(false, 'Exit After N Bars');
    }
    html += yesNo(code.has_exit_signals, 'Signal-Based Exit');
    html += '</div></div>';

    html += '<div class="ov-section"><div class="ov-section-title">Exit Methods</div><div class="sp-stats">';
    html += `<div class="sp-row"><span class="sp-val" style="color:var(--text-secondary);font-size:0.78rem">${code.exit_summary}</span></div>`;
    html += '</div></div>';

    if (code.indicators && code.indicators.length) {
      html += '<div class="ov-section"><div class="ov-section-title">Indicators</div>';
      html += '<div class="code-indicators">';
      code.indicators.forEach(ind => { html += `<span class="code-ind-tag">${ind}</span>`; });
      html += '</div></div>';
    }
    if (code.entry_refs && code.entry_refs.length) {
      html += '<div class="ov-section"><div class="ov-section-title">Entry Reference Levels</div>';
      html += '<div class="code-indicators">';
      code.entry_refs.forEach(ref => { html += `<span class="code-ref-tag">${ref}</span>`; });
      html += '</div></div>';
    }

    html += '<div class="ov-section"><div class="ov-section-title">Trading Options</div><div class="sp-stats">';
    if (code.time_filter) {
      html += `<div class="sp-row"><span class="sp-label">Time Filter</span><span class="sp-val">${code.time_filter}</span></div>`;
    }
    if (code.order_valid_bars) {
      html += `<div class="sp-row"><span class="sp-label">Order Valid For</span><span class="sp-val">${code.order_valid_bars} bars</span></div>`;
    }
    html += '</div></div>';

    html += '</div>';
    body.innerHTML = html;
    sp.classList.remove('hidden');
    $$('.analyze-btn').forEach(b => b.classList.toggle('active',
      b.dataset.name === name && b.dataset.action === 'code'));
  }

  // === Export Strategy Data ===
  function exportStrategy(name) {
    const ranking = DATA.ranking.find(r => r.name === name);
    const summary = DATA.summary.find(s => s.name === name);
    const portfolio = DATA.portfolio.find(p => p.name === name);
    const overview = DATA.overviews[name] || null;
    const code = DATA.strategy_codes[name] || null;
    const sqxMeta = DATA.sqx_metadata[name] || null;

    const exportData = {
      _export_version: '1.0',
      _exported_at: new Date().toISOString(),
      _source: 'strategy_correlation_analysis',
      strategy_name: name,

      // ── Maps to: backtest_data TEXT column ──
      backtest_data: {
        ranking: ranking ? {
          composite_score: ranking.score,
          net_profit: ranking.net_profit,
          ret_dd: ranking.ret_dd,
          mc95_ret_dd: ranking.mc95_ret_dd,
          mc95_ret_dd_tick: ranking.mc95_ret_dd_tick,
          wl_ratio: ranking.wl_ratio,
          profit_factor: ranking.pf,
          sharpe: ranking.sharpe,
          recovery: ranking.recovery,
          lr_correlation: ranking.lr_corr,
          win_rate: ranking.win_rate,
          total_trades: ranking.trades,
          max_dd_dollars: ranking.dd_dollar,
          max_dd_pct: ranking.dd_pct,
          rank: ranking.rank,
        } : null,
        portfolio: portfolio ? {
          decision: portfolio.decision,
          cluster: portfolio.cluster,
          reason: portfolio.reason,
          direction: portfolio.direction,
          total_pnl: portfolio.total_pnl,
          avg_trade: portfolio.avg_trade,
        } : null,
        overview: overview || null,
        strategy_profile: code ? {
          direction: code.direction,
          style: code.style,
          entry_type: code.entry_type,
          indicators: code.indicators,
          entry_refs: code.entry_refs,
          trailing_stop: code.trailing_stop,
          ts_activation: code.ts_activation,
          move_sl_be: code.move_sl_be,
          profit_target: code.profit_target,
          exit_after_bars: code.exit_after_bars,
          has_exit_signals: code.has_exit_signals,
          time_filter: code.time_filter,
          order_valid_bars: code.order_valid_bars,
          exit_summary: code.exit_summary,
        } : null,
        sqx_metadata: sqxMeta ? {
          complexity: sqxMeta.complexity,
        } : null,
        summary: summary ? {
          total_pnl: summary.total_pnl,
          num_trades: summary.num_trades,
          win_rate: summary.win_rate,
          avg_trade: summary.avg_trade,
          direction: summary.direction,
        } : null,
      },

      // ── Maps to: backtest_equitychart TEXT column ──
      backtest_equitychart: (ranking ? ranking.chart : (summary ? summary.chart : null)) || null,

      // ── Maps to: pseudo_code TEXT column ──
      pseudo_code: (code ? code.raw_content : null) || null,
    };

    downloadExport(name, exportData);
  }

  function downloadExport(name, data) {
    const json = JSON.stringify(data, null, 2);
    const blob = new Blob([json], {type: 'application/json'});
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = name.replace(/[^a-zA-Z0-9._()-]/g, '_') + '_export.json';
    document.body.appendChild(a);
    a.click();
    document.body.removeChild(a);
    URL.revokeObjectURL(url);
  }

  // === Chart Lightbox (click to enlarge) ===
  function openLightbox(src, title) {
    const lb = $('#chartLightbox');
    $('#lbImg').src = src;
    $('#lbTitle').textContent = title;
    lb.classList.add('active');
  }
  function closeLightbox() {
    $('#chartLightbox').classList.remove('active');
  }
  // Make openLightbox available to inline onclick
  window.openLightbox = openLightbox;
  $('#chartLightbox').addEventListener('click', e => {
    if (e.target === $('#chartLightbox') || e.target === $('#lbClose')) closeLightbox();
  });
  document.addEventListener('keydown', e => {
    if (e.key === 'Escape') { closeLightbox(); closeSidePanel(); }
  });

  // === Sortable Table Helper ===
  function renderSortableTable(containerId, headers, rows, options = {}) {
    const container = $(containerId);
    let sortCol = options.defaultSort || 0;
    let sortDir = options.defaultDir || 'asc';

    function render() {
      const sorted = [...rows].sort((a, b) => {
        let va = a[sortCol], vb = b[sortCol];
        if (typeof va === 'string') {
          // Try parsing as number after stripping $, %, commas
          const na = parseFloat(String(va).replace(/[$,%\s]/g, '').replace(/,/g, ''));
          const nb = parseFloat(String(vb).replace(/[$,%\s]/g, '').replace(/,/g, ''));
          if (!isNaN(na) && !isNaN(nb)) {
            return sortDir === 'asc' ? na - nb : nb - na;
          }
          va = va.toLowerCase(); vb = vb.toLowerCase();
          return sortDir === 'asc' ? va.localeCompare(vb) : vb.localeCompare(va);
        }
        return sortDir === 'asc' ? va - vb : vb - va;
      });

      let html = '<div class="table-wrap"><table><thead><tr>';
      headers.forEach((h, i) => {
        const cls = (i === sortCol) ? `sorted-${sortDir}` : '';
        const noSort = h.noSort ? ' no-sort' : '';
        html += `<th class="${cls}${noSort}" data-col="${i}">${h.label}</th>`;
      });
      html += '</tr></thead><tbody>';
      sorted.forEach(row => {
        const tierClass = row._tier || '';
        const abandonClass = row._abandon ? ' abandon-row' : '';
        html += `<tr class="${tierClass}${abandonClass}">`;
        headers.forEach((h, i) => {
          const cls = h.text ? ' class="text-col"' : '';
          if (h.rawHtml) {
            html += `<td${cls}>${row[i]}</td>`;
          } else {
            html += `<td${cls}>${row[i]}</td>`;
          }
        });
        html += '</tr>';
      });
      html += '</tbody></table></div>';
      container.innerHTML = html;

      container.querySelectorAll('th:not(.no-sort)').forEach(th => {
        th.addEventListener('click', () => {
          const col = +th.dataset.col;
          if (col === sortCol) {
            sortDir = sortDir === 'asc' ? 'desc' : 'asc';
          } else {
            sortCol = col; sortDir = 'asc';
          }
          render();
        });
      });
    }
    render();
  }

  // === Rankings Panel ===
  (function buildRankings() {
    const el = $('#panel-ranking');
    if (!DATA.ranking.length) {
      el.innerHTML = '<div class="section-header">MT5 Backtest Rankings</div><p style="color:var(--text-muted)">No MT5 reports found — ranking not available. Place MT5 .htm reports alongside your trade CSVs.</p>';
      return;
    }
    el.innerHTML = '<div class="section-header">MT5 Backtest Rankings</div><div class="section-sub">' +
      DATA.ranking.length + ' strategies ranked by multi-metric composite score</div><div id="rankingTable"></div>';

    // Build lookup for portfolio decisions
    const portfolioDecisions = {};
    if (DATA.portfolio) {
      DATA.portfolio.forEach(p => { portfolioDecisions[p.name] = p.decision; });
    }

    const headers = [
      {label:'#'}, {label:'Strategy', text:true, rawHtml:true}, {label:'Score'}, {label:'Net Profit'},
      {label:'Ret/DD'}, {label:'MC95 Ret/DD'}, {label:'MC95 Tick'}, {label:'W/L'}, {label:'PF'}, {label:'Sharpe'},
      {label:'Recovery'}, {label:'LR Corr'}, {label:'Win%'}, {label:'Trades'},
      {label:'DD ($)'}, {label:'DD %'}, {label:'Analyze', noSort:true, rawHtml:true}
    ];
    const rows = DATA.ranking.map(r => {
      const hasOv = !!DATA.overviews[r.name];
      const hasCode = !!DATA.strategy_codes[r.name];
      const ovBtn = hasOv
        ? `<button class="analyze-btn ov-btn" data-name="${r.name}" data-action="overview">Overview</button>`
        : `<button class="analyze-btn ov-btn disabled">Overview</button>`;
      const codeBtn = hasCode
        ? `<button class="analyze-btn code-btn" data-name="${r.name}" data-action="code">Strategy</button>`
        : `<button class="analyze-btn code-btn disabled">Strategy</button>`;
      const chartBtn = r.chart
        ? `<button class="analyze-btn chart-btn" data-name="${r.name}" data-chart="${r.chart}" data-action="chart">Chart</button>`
        : `<button class="analyze-btn chart-btn disabled">Chart</button>`;
      const exportBtn = `<button class="analyze-btn export-btn" data-name="${r.name}" data-action="export">Export</button>`;

      // Show strategy name with ABANDON badge if applicable
      const decision = portfolioDecisions[r.name];
      const nameDisplay = decision === 'ABANDON'
        ? `<span class="abandon-name">${r.name}</span> <span class="tag tag-abandon" style="font-size:10px;padding:1px 6px;vertical-align:middle;">✗ ABANDON</span>`
        : r.name;

      const row = [
        r.rank, nameDisplay, r.score, '$' + r.net_profit.toLocaleString(undefined, {minimumFractionDigits:2}),
        r.ret_dd, r.mc95_ret_dd || '—', r.mc95_ret_dd_tick || '—', r.wl_ratio, r.pf, r.sharpe,
        r.recovery, r.lr_corr, r.win_rate + '%', r.trades,
        '$' + r.dd_dollar.toLocaleString(undefined, {minimumFractionDigits:2}), r.dd_pct + '%',
        `<div class="btn-group">${ovBtn}${codeBtn}${chartBtn}${exportBtn}</div>`
      ];
      row._tier = r.rank <= 5 ? 'tier-1' : r.rank <= 10 ? 'tier-2' : r.rank <= 25 ? 'tier-3' : '';
      if (decision === 'ABANDON') row._abandon = true;
      return row;
    });
    renderSortableTable('#rankingTable', headers, rows, {defaultSort: 0, defaultDir: 'asc'});

    // Button delegation
    el.addEventListener('click', e => {
      const btn = e.target.closest('.analyze-btn:not(.disabled)');
      if (!btn) return;
      const name = btn.dataset.name;
      const action = btn.dataset.action;

      if (action === 'overview') {
        if (activeSideStrategy === name + '_ov') { closeSidePanel(); return; }
        activeSideStrategy = name + '_ov';
        openOverviewPanel(name);
      } else if (action === 'code') {
        if (activeSideStrategy === name + '_code') { closeSidePanel(); return; }
        openCodePanel(name);
      } else if (action === 'chart') {
        if (activeSideStrategy === name) { closeSidePanel(); return; }
        const r = DATA.ranking.find(x => x.name === name);
        const stats = [
          ['Composite Score', r.score], ['Net Profit', '$' + r.net_profit.toLocaleString()],
          ['Ret/DD Ratio', r.ret_dd], ['MC95 Ret/DD', r.mc95_ret_dd || '—'],
          ['MC95 Ret/DD Tick', r.mc95_ret_dd_tick || '—'], ['Profit Factor', r.pf],
          ['Sharpe Ratio', r.sharpe], ['Recovery Factor', r.recovery],
          ['LR Correlation', r.lr_corr], ['Win Rate', r.win_rate + '%'],
          ['Total Trades', r.trades], ['Max DD', '$' + r.dd_dollar.toLocaleString() + ' (' + r.dd_pct + '%)'],
        ];
        openSidePanel(name, btn.dataset.chart, stats);
      } else if (action === 'export') {
        exportStrategy(name);
      }
    });
  })();

  // === MC Failed Panel ===
  (function buildMCFailed() {
    const el = $('#panel-mc-failed');
    if (!DATA.mc_failed || !DATA.mc_failed.length) {
      el.innerHTML = '<div class="section-header">MC Failed Strategies</div><p style="color:var(--text-muted)">No strategies failed the Monte Carlo threshold, or no MC results were provided.</p>';
      return;
    }
    el.innerHTML = '<div class="section-header">MC Failed Strategies</div><div class="section-sub">' +
      DATA.mc_failed.length + ' strategies failed MC95 Ret/DD threshold — excluded from analysis</div><div id="mcFailedTable"></div>';

    const headers = [
      {label:'Strategy', text:true}, {label:'Orig Net Profit'}, {label:'MC95 Net Profit'},
      {label:'Orig Ret/DD'}, {label:'MC95 Ret/DD'}, {label:'MC95 Max DD'},
      {label:'MC95 DD %'}, {label:'Trades'}
    ];
    const rows = DATA.mc_failed.map(r => {
      return [
        r.name,
        '$' + r.orig_net_profit.toLocaleString(undefined, {minimumFractionDigits:2}),
        '$' + r.mc95_net_profit.toLocaleString(undefined, {minimumFractionDigits:2}),
        r.orig_ret_dd, r.mc95_ret_dd,
        '$' + r.mc95_max_dd.toLocaleString(undefined, {minimumFractionDigits:2}),
        r.mc95_max_dd_pct + '%', r.mc95_trades
      ];
    });
    renderSortableTable('#mcFailedTable', headers, rows, {defaultSort: 4, defaultDir: 'desc'});
  })();

  // === Portfolio Panel ===
  (function buildPortfolio() {
    const el = $('#panel-portfolio');
    el.innerHTML = '<div class="section-header">Portfolio Selection</div><div class="section-sub">Best composite score from each correlation cluster</div><div id="portfolioTable"></div>';

    const headers = [
      {label:'Strategy', text:true}, {label:'Decision', rawHtml:true}, {label:'Cluster'},
      {label:'Reason', text:true}, {label:'Direction', rawHtml:true},
      {label:'Total P&L'}, {label:'Avg Trade'}, {label:'Win%'},
      {label:'Analyze', noSort:true, rawHtml:true}
    ];
    const rows = DATA.portfolio.map(p => {
      const decTag = p.decision === 'KEEP'
        ? '<span class="tag tag-keep">✓ KEEP</span>'
        : '<span class="tag tag-abandon">✗ ABANDON</span>';
      const dirCls = p.direction === 'Long Only' ? 'tag-long' : p.direction === 'Short Only' ? 'tag-short' : 'tag-both';
      const hasOv = !!DATA.overviews[p.name];
      const hasCode = !!DATA.strategy_codes[p.name];
      const ovBtn = hasOv
        ? `<button class="analyze-btn ov-btn" data-name="${p.name}" data-action="overview">Overview</button>`
        : `<button class="analyze-btn ov-btn disabled">Overview</button>`;
      const codeBtn = hasCode
        ? `<button class="analyze-btn code-btn" data-name="${p.name}" data-action="code">Strategy</button>`
        : `<button class="analyze-btn code-btn disabled">Strategy</button>`;
      const chartBtn = p.chart
        ? `<button class="analyze-btn chart-btn" data-name="${p.name}" data-chart="${p.chart}" data-action="chart">Chart</button>`
        : `<button class="analyze-btn chart-btn disabled">Chart</button>`;
      const exportBtn = `<button class="analyze-btn export-btn" data-name="${p.name}" data-action="export">Export</button>`;
      return [
        p.name, decTag, p.cluster, p.reason,
        `<span class="${dirCls}">${p.direction}</span>`,
        '$' + p.total_pnl.toLocaleString(undefined, {minimumFractionDigits:2}),
        '$' + p.avg_trade.toFixed(2), p.win_rate + '%',
        `<div class="btn-group">${ovBtn}${codeBtn}${chartBtn}${exportBtn}</div>`
      ];
    });
    renderSortableTable('#portfolioTable', headers, rows, {defaultSort: 2, defaultDir: 'asc'});

    el.addEventListener('click', e => {
      const btn = e.target.closest('.analyze-btn:not(.disabled)');
      if (!btn) return;
      const name = btn.dataset.name;
      const action = btn.dataset.action;

      if (action === 'overview') {
        if (activeSideStrategy === name + '_ov') { closeSidePanel(); return; }
        activeSideStrategy = name + '_ov';
        openOverviewPanel(name);
      } else if (action === 'code') {
        if (activeSideStrategy === name + '_code') { closeSidePanel(); return; }
        openCodePanel(name);
      } else if (action === 'chart') {
        if (activeSideStrategy === name) { closeSidePanel(); return; }
        const p = DATA.portfolio.find(x => x.name === name);
        const stats = [
          ['Decision', p.decision], ['Cluster', p.cluster], ['Direction', p.direction],
          ['Total P&L', '$' + p.total_pnl.toLocaleString()],
          ['Avg Trade', '$' + p.avg_trade.toFixed(2)], ['Win Rate', p.win_rate + '%'],
        ];
        openSidePanel(name, btn.dataset.chart, stats);
      } else if (action === 'export') {
        exportStrategy(name);
      }
    });
  })();

  // === Correlation Panel ===
  (function buildCorrelation() {
    const el = $('#panel-correlation');
    let currentFreq = 'weekly';

    function render() {
      const matrix = DATA.correlations[currentFreq];
      const names = DATA.names;
      const n = names.length;

      // Truncate names for headers
      const short = names.map(nm => {
        const parts = nm.split(' ');
        return parts.length > 3 ? parts.slice(-2).join('.') : nm;
      });

      let html = '<div class="section-header">Correlation Matrix</div>';
      html += '<div class="heatmap-controls">';
      ['daily','weekly','monthly'].forEach(f => {
        html += `<button class="hm-btn${f===currentFreq?' active':''}" data-freq="${f}">${f.charAt(0).toUpperCase()+f.slice(1)}</button>`;
      });
      html += '</div>';

      html += `<div class="heatmap-grid" style="grid-template-columns: 120px repeat(${n}, 1fr);">`;
      // Header row
      html += '<div></div>';
      short.forEach((nm, i) => {
        html += `<div class="hm-header" title="${names[i]}">${nm}</div>`;
      });
      // Data rows
      names.forEach((n1, i) => {
        html += `<div class="hm-row-header" title="${n1}">${short[i]}</div>`;
        names.forEach((n2, j) => {
          const val = matrix[i][j];
          const abs = Math.abs(val);
          let bg;
          if (i === j) bg = 'var(--bg-active)';
          else if (abs >= 0.7) bg = 'rgba(248,113,113,0.35)';
          else if (abs >= 0.5) bg = 'rgba(251,191,36,0.30)';
          else if (abs >= 0.3) bg = 'rgba(74,222,128,0.20)';
          else bg = 'rgba(74,222,128,0.08)';
          const textColor = i === j ? 'var(--text-muted)' : 'var(--text-primary)';
          html += `<div class="hm-cell" style="background:${bg};color:${textColor}" title="${names[i]} vs ${names[j]}: ${val}">${val.toFixed(2)}</div>`;
        });
      });
      html += '</div>';

      html += '<div class="colour-key">';
      html += '<span><span class="ck-swatch" style="background:rgba(74,222,128,0.15)"></span>&lt; 0.3 Low</span>';
      html += '<span><span class="ck-swatch" style="background:rgba(74,222,128,0.35)"></span>0.3–0.5</span>';
      html += '<span><span class="ck-swatch" style="background:rgba(251,191,36,0.45)"></span>0.5–0.7</span>';
      html += '<span><span class="ck-swatch" style="background:rgba(248,113,113,0.5)"></span>&gt; 0.7 High</span>';
      html += '</div>';

      el.innerHTML = html;

      el.querySelectorAll('.hm-btn').forEach(btn => {
        btn.addEventListener('click', () => { currentFreq = btn.dataset.freq; render(); });
      });
    }
    render();
  })();

  // === Overlap Panel ===
  (function buildOverlap() {
    const el = $('#panel-overlap');
    el.innerHTML = '<div class="section-header">Trade Overlap Analysis</div><div class="section-sub">Pairwise time-in-market overlap between strategies</div><div id="overlapTable"></div>';

    const headers = [
      {label:'Pair', text:true}, {label:'Overlapping'}, {label:'Overlap %'},
      {label:'Same Dir'}, {label:'Opp Dir'}, {label:'Assessment', rawHtml:true}
    ];
    const rows = DATA.overlap.map(o => {
      let tag, color;
      if (o.pct >= 80) { tag = 'Very High'; color = 'var(--red)'; }
      else if (o.pct >= 60) { tag = 'High'; color = 'var(--amber)'; }
      else if (o.pct >= 40) { tag = 'Moderate'; color = 'var(--text-secondary)'; }
      else { tag = 'Low — Good'; color = 'var(--green)'; }
      return [o.pair, `${o.overlaps}/${o.total}`, o.pct + '%', o.same, o.opp,
              `<span style="color:${color};font-weight:500">${tag}</span>`];
    });
    renderSortableTable('#overlapTable', headers, rows, {defaultSort: 2, defaultDir: 'desc'});
  })();

  // === Clusters Panel ===
  (function buildClusters() {
    const el = $('#panel-clusters');
    const bestInCluster = {};
    DATA.portfolio.filter(p => p.decision === 'KEEP').forEach(p => { bestInCluster[p.cluster] = p.name; });

    let html = '<div class="section-header">Correlation Clusters</div>';
    html += `<div class="section-sub">${DATA.clusters.length} independent clusters identified — pick the best performer from each</div>`;
    html += '<div class="cluster-grid">';
    DATA.clusters.forEach(cl => {
      const best = bestInCluster[cl.id] || '';
      html += '<div class="cluster-card">';
      html += `<h4>Cluster ${cl.id} <span style="color:var(--text-muted);font-weight:400;font-size:0.75rem">(${cl.count} strateg${cl.count===1?'y':'ies'})</span></h4>`;
      cl.members.forEach(m => {
        const isBest = m === best;
        html += `<div class="cluster-member${isBest ? ' is-best' : ''}">${isBest ? '★ ' : ''}${m}</div>`;
      });
      html += '</div>';
    });
    html += '</div>';

    // Best pairs sub-section
    html += '<div style="margin-top:28px"><div class="section-header">Best Diversification Pairs</div>';
    html += '<div class="section-sub">Lowest weekly correlation — best candidates for combining</div>';
    html += '<div id="bestPairsTable"></div></div>';

    el.innerHTML = html;

    const bpHeaders = [{label:'#'}, {label:'Pair', text:true}, {label:'Weekly Corr'}, {label:'Daily Corr'}];
    const bpRows = DATA.best_pairs.slice(0, 20).map((p, i) => [i+1, p.pair, p.weekly, p.daily]);
    renderSortableTable('#bestPairsTable', bpHeaders, bpRows, {defaultSort: 0, defaultDir: 'asc'});
  })();

  // === Drawdowns Panel ===
  (function buildDrawdowns() {
    const el = $('#panel-drawdowns');
    el.innerHTML = '<div class="section-header">Max Drawdown Periods</div><div class="section-sub">Worst drawdown period for each strategy</div><div id="ddTable"></div>';

    const headers = [
      {label:'Strategy', text:true}, {label:'Max DD ($)'}, {label:'DD Start', text:true},
      {label:'DD End', text:true}, {label:'Duration (days)'},
      {label:'Analyze', noSort:true, rawHtml:true}
    ];
    const rows = DATA.drawdowns.map(d => {
      const s = DATA.summary.find(x => x.name === d.name);
      const hasOv = !!DATA.overviews[d.name];
      const hasCode = !!DATA.strategy_codes[d.name];
      const ovBtn = hasOv
        ? `<button class="analyze-btn ov-btn" data-name="${d.name}" data-action="overview">Overview</button>`
        : `<button class="analyze-btn ov-btn disabled">Overview</button>`;
      const codeBtn = hasCode
        ? `<button class="analyze-btn code-btn" data-name="${d.name}" data-action="code">Strategy</button>`
        : `<button class="analyze-btn code-btn disabled">Strategy</button>`;
      const chartBtn = s && s.chart
        ? `<button class="analyze-btn chart-btn" data-name="${d.name}" data-chart="${s.chart}" data-action="chart">Chart</button>`
        : `<button class="analyze-btn chart-btn disabled">Chart</button>`;
      const exportBtn = `<button class="analyze-btn export-btn" data-name="${d.name}" data-action="export">Export</button>`;
      return [d.name, '$' + d.max_dd.toLocaleString(undefined,{minimumFractionDigits:2}), d.dd_start, d.dd_end, d.dd_days,
        `<div class="btn-group">${ovBtn}${codeBtn}${chartBtn}${exportBtn}</div>`];
    });
    renderSortableTable('#ddTable', headers, rows, {defaultSort: 1, defaultDir: 'asc'});

    el.addEventListener('click', e => {
      const btn = e.target.closest('.analyze-btn:not(.disabled)');
      if (!btn) return;
      const name = btn.dataset.name;
      const action = btn.dataset.action;

      if (action === 'overview') {
        if (activeSideStrategy === name + '_ov') { closeSidePanel(); return; }
        activeSideStrategy = name + '_ov';
        openOverviewPanel(name);
      } else if (action === 'code') {
        if (activeSideStrategy === name + '_code') { closeSidePanel(); return; }
        openCodePanel(name);
      } else if (action === 'chart') {
        if (activeSideStrategy === name) { closeSidePanel(); return; }
        const d = DATA.drawdowns.find(x => x.name === name);
        const s = DATA.summary.find(x => x.name === name);
        const stats = [
          ['Max Drawdown', '$' + d.max_dd.toLocaleString()],
          ['DD Start', d.dd_start], ['DD End', d.dd_end],
          ['Duration', d.dd_days + ' days'],
          ['Total P&L', '$' + (s ? s.total_pnl.toLocaleString() : '—')],
        ];
        openSidePanel(name, btn.dataset.chart, stats);
      } else if (action === 'export') {
        exportStrategy(name);
      }
    });
  })();

  // === Methodology Panel ===
  (function buildMethodology() {
    const el = $('#panel-methodology');
    const weights = [
      ['Net Profit', '15%', 'Absolute profitability — ensures the strategy generates meaningful returns'],
      ['Ret/DD Ratio', '15%', 'Net profit / max equity drawdown — risk-adjusted return quality'],
      ['Profit Factor', '10%', 'Gross profit / gross loss — measures overall reward-to-risk'],
      ['Sharpe Ratio', '10%', 'Risk-adjusted return considering volatility — penalises inconsistent curves'],
      ['Recovery Factor', '10%', 'Net profit / max drawdown — how quickly the strategy recovers from losses'],
      ['LR Correlation', '10%', 'Linear regression R of equity curve — measures equity curve smoothness/stability'],
      ['Win/Loss Ratio', '5%', 'Average win / average loss — reward-to-risk per trade'],
      ['# Trades', '5%', 'Statistical significance — more trades = more confidence in the edge'],
      ['Expected Payoff', '5%', 'Average P&L per trade — how much edge each trade provides'],
      ['Win Rate %', '5%', 'Percentage of winning trades — consistency of the edge'],
      ['Max DD %', '5%', 'Inverted — lower max drawdown percentage is better'],
      ['LR Std Error', '5%', 'Inverted — lower standard error = smoother equity curve'],
    ];

    el.innerHTML = `
      <div class="section-header">MT5 Backtest Ranking Methodology</div>

      <div class="meth-section">
        <div class="meth-heading">Overview</div>
        <p>Strategies that pass Monte Carlo testing in StrategyQuant X are validated by running backtests in MetaTrader 5. The MT5 HTML reports are then parsed to extract performance metrics, which are used to rank strategies using a percentile-based composite scoring methodology.</p>
      </div>

      <div class="meth-section">
        <div class="meth-heading">Scoring Method</div>
        <p>Each metric is converted to a percentile rank (0–1), then weighted and summed. Metrics where lower is better (e.g. drawdown %) are inverted before weighting. This ensures strategies must be consistently good across multiple dimensions.</p>
      </div>

      <div class="meth-section">
        <div class="meth-heading">Metric Weights</div>
        <table class="meth-table">
          <thead><tr><th>Metric</th><th>Weight</th><th>Rationale</th></tr></thead>
          <tbody>${weights.map(([m,w,r]) => `<tr><td>${m}</td><td style="text-align:center">${w}</td><td>${r}</td></tr>`).join('')}</tbody>
        </table>
      </div>

      <div class="meth-section">
        <div class="meth-heading">Colour Coding</div>
        <div class="meth-colours">
          <span class="meth-swatch tier-1-bg"></span> Top 5 &nbsp;&nbsp;
          <span class="meth-swatch tier-2-bg"></span> Top 10 &nbsp;&nbsp;
          <span class="meth-swatch tier-3-bg"></span> Top 25
        </div>
      </div>

      <div class="meth-section">
        <div class="meth-heading">Data Source</div>
        <ul class="meth-notes">
          <li>Metrics are parsed directly from MT5 Strategy Tester HTML reports</li>
          <li>Reports must be in the same folder as the trade CSV files (or specified via <code>--mt5-reports</code>)</li>
          <li>Strategy names are matched by normalising filenames (underscores / spaces / dots)</li>
          <li>Ret/DD Ratio and Win/Loss Ratio are computed from the parsed MT5 values</li>
          <li>Weights are adjustable in the <code>MT5_HIGHER_IS_BETTER</code> / <code>MT5_LOWER_IS_BETTER</code> config at top of script</li>
        </ul>
      </div>
    `;
  })();

})();
"""


# ============================================================================
# MAIN
# ============================================================================
def main():
    init_colors()
    
    folder_path = None
    mt5_folder = None
    mc_results_path = None
    mc95_threshold = DEFAULT_MC95_THRESHOLD

    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == '--mt5-reports' and i + 1 < len(args):
            mt5_folder = args[i + 1]
            i += 2
        elif args[i] == '--mc-results' and i + 1 < len(args):
            mc_results_path = args[i + 1]
            i += 2
        elif args[i] == '--mc95-threshold' and i + 1 < len(args):
            mc95_threshold = float(args[i + 1])
            i += 2
        elif folder_path is None:
            folder_path = args[i]
            i += 1
        else:
            i += 1

    if folder_path is None:
        folder_path = os.getcwd()
    folder_path = os.path.abspath(folder_path)

    if mt5_folder is None:
        mt5_folder = folder_path
    else:
        mt5_folder = os.path.abspath(mt5_folder)

    # Record start time
    start_time = time.time()
    start_datetime = datetime.now()

    print(f"\n{Colors.CYAN}{'=' * 60}{Colors.RESET}")
    print(f"{Colors.CYAN}Strategy Correlation Analysis{Colors.RESET}")
    print(f"{Colors.CYAN}{'=' * 60}{Colors.RESET}")
    print()
    print(f"{Colors.CYAN}Started:{Colors.RESET} {Colors.GREEN}{start_datetime.strftime('%Y-%m-%d %H:%M:%S')}{Colors.RESET}")
    print()
    print(f"Scanning: {folder_path}")
    if mt5_folder != folder_path:
        print(f"MT5 reports: {mt5_folder}")

    # Step 1: Load Monte Carlo results and filter strategies
    mc95_data = {}
    mc_passing_names = None  # None = no MC filter (analyse all CSVs)
    mc_failed = []

    if mc_results_path:
        mc_results_path = os.path.abspath(mc_results_path)
        print(f"\nLoading Monte Carlo results: {mc_results_path}")
        mc_passing_names, mc95_data, mc_failed = load_mc_results(mc_results_path, mc95_threshold)
        if not mc_passing_names:
            print("ERROR: No strategies passed MC95 threshold - nothing to analyse.")
            sys.exit(1)
        print(f"  -> {len(mc_passing_names)} strategies passed MC filter")
    else:
        print("\n  No --mc-results provided - analysing all trade CSVs")

    print()

    # Step 2: Load trade CSVs (filtered to MC-passing strategies if applicable)
    strategies = load_strategies(folder_path)

    # Filter to only MC-passing strategies
    if mc_passing_names is not None:
        filtered = {}
        mc_name_set = set(mc_passing_names)
        for csv_name, df in strategies.items():
            matched = match_csv_to_mc(csv_name, mc_name_set)
            if matched:
                filtered[csv_name] = df
            else:
                print(f"  SKIPPING (not in MC pass list): {csv_name}")

        if len(filtered) == 0:
            print(f"ERROR: No strategies matched MC results.")
            print("  Check that trade CSV names match BatchMC_Results.csv strategy names.")
            sys.exit(1)
        
        if len(filtered) == 1:
            print(f"\n{Colors.YELLOW}WARNING: Only 1 strategy matched MC results - skipping correlation analysis.{Colors.RESET}")
            skip_correlation = True
        else:
            skip_correlation = False

        print(f"\n  {len(filtered)}/{len(strategies)} strategies matched MC pass list")
        strategies = filtered
    else:
        # No MC filter - check if we have enough strategies for correlation
        skip_correlation = len(strategies) < 2
        if skip_correlation and len(strategies) == 1:
            print(f"\n{Colors.YELLOW}WARNING: Only 1 strategy found - skipping correlation analysis.{Colors.RESET}")
        elif len(strategies) == 0:
            print(f"ERROR: No strategies found to analyse.")
            sys.exit(1)

    names = sorted(strategies.keys())
    n = len(names)
    print(f"\nAnalysing {n} strategies.\n")

    print("Computing strategy statistics...")
    stats = compute_strategy_stats(strategies)

    # Initialize correlation and overlap data structures
    corr_daily = pd.DataFrame()
    corr_weekly = pd.DataFrame()
    corr_monthly = pd.DataFrame()
    overlap_data = {}
    clusters = [[name] for name in names]  # Each strategy in its own cluster by default

    if not skip_correlation:
        print("Computing P&L correlations (daily, weekly, monthly)...")
        print("  (excluding periods where both strategies had no activity)")
        daily_df = build_pnl_series(strategies, 'D')
        weekly_df = build_pnl_series(strategies, 'W')
        monthly_df = build_pnl_series(strategies, 'M')
        corr_daily = compute_pairwise_correlation(daily_df)
        corr_weekly = compute_pairwise_correlation(weekly_df)
        corr_monthly = compute_pairwise_correlation(monthly_df, min_observations=6)

        print("Computing pairwise trade overlap (this may take a while for many strategies)...")
        total_pairs = n * (n - 1) // 2
        pair_count = 0
        for i, n1 in enumerate(names):
            for j in range(i + 1, len(names)):
                n2 = names[j]
                pair_count += 1
                print(f"  [{pair_count}/{total_pairs}] {n1} vs {n2}...", end='', flush=True)
                overlaps, same, opp = compute_trade_overlap(strategies[n1], strategies[n2])
                total1 = len(strategies[n1])
                pct = overlaps / total1 if total1 > 0 else 0
                overlap_data[(n1, n2)] = {
                    'overlaps': overlaps, 'total': total1,
                    'pct': pct, 'same': same, 'opp': opp
                }
                print(f" {pct:.1%}")

        print("\nIdentifying correlation clusters...")
        clusters = identify_clusters(names, corr_weekly, threshold=0.5)
        for i, cluster in enumerate(clusters):
            print(f"  Cluster {i + 1}: {', '.join(cluster)}")
    else:
        print(f"{Colors.YELLOW}Skipping correlation analysis (need at least 2 strategies){Colors.RESET}")
        # Create placeholder 1x1 correlation matrices for single strategy
        if n == 1:
            corr_daily = pd.DataFrame([[1.0]], index=names, columns=names)
            corr_weekly = pd.DataFrame([[1.0]], index=names, columns=names)
            corr_monthly = pd.DataFrame([[1.0]], index=names, columns=names)

    print("\nGenerating Excel report + HTML dashboard...")
    output_path, keep, abandon, strategy_scores = generate_report(
        folder_path, strategies, stats, names,
        corr_daily, corr_weekly, corr_monthly,
        overlap_data, clusters, mt5_folder=mt5_folder,
        mc95_data=mc95_data, mc_failed=mc_failed,
    )

    print(f"\n{Colors.CYAN}{'=' * 60}{Colors.RESET}")
    print(f"{Colors.GREEN}DONE! Report saved to:{Colors.RESET}")
    print(f"  {output_path}")
    print(f"{Colors.CYAN}{'=' * 60}{Colors.RESET}")

    selection_method = 'composite score' if strategy_scores else 'P&L'
    print(f"\n{Colors.CYAN}Quick Summary:{Colors.RESET}")
    print(f"  Strategies analysed: {n}")
    if mc_results_path:
        print(f"  MC95 Ret/DD threshold: >= {mc95_threshold}")
    
    if not skip_correlation:
        print(f"  Correlation clusters: {len(clusters)}")
        print(f"  Suggested portfolio size: ~{len(clusters)} strategies (1 per cluster)")
    else:
        print(f"  Correlation analysis: {Colors.YELLOW}SKIPPED (insufficient strategies){Colors.RESET}")
    
    print(f"  Selection method: {selection_method}")

    print(f"\n  + KEEP ({len(keep)}):")
    for s in sorted(keep.keys(), key=lambda x: keep[x]):
        csize = len(clusters[keep[s] - 1]) if keep[s] - 1 < len(clusters) else 1
        score_tag = f'  [Score: {strategy_scores[s]:.3f}]' if strategy_scores and s in strategy_scores else ''
        tag = '' if csize == 1 else f' (best of {csize} in cluster)'
        mc_tag = ''
        matched = match_csv_to_mc(s, set(mc95_data.keys()))
        if matched and matched in mc95_data:
            mc_tag = f'  [MC95 Ret/DD: {mc95_data[matched]["mc95_ret_dd"]:.2f}]'
        if skip_correlation:
            print(f"    {s} - ${stats[s]['total_pnl']:,.0f}{score_tag}{mc_tag}")
        else:
            print(f"    Cluster {keep[s]}: {s} - ${stats[s]['total_pnl']:,.0f}{tag}{score_tag}{mc_tag}")

    if abandon:
        print(f"\n  x ABANDON ({len(abandon)}):")
        for s in sorted(abandon.keys(), key=lambda x: abandon[x][0]):
            cid, replaced = abandon[s]
            score_tag = f'  [Score: {strategy_scores[s]:.3f}]' if strategy_scores and s in strategy_scores else ''
            print(f"    Cluster {cid}: {s} - redundant with {replaced}{score_tag}")

    if not skip_correlation:
        high_corr_pairs = []
        for i, n1 in enumerate(names):
            for j in range(i + 1, len(names)):
                n2 = names[j]
                wc = corr_weekly.loc[n1, n2]
                if abs(wc) >= CORR_HIGH:
                    high_corr_pairs.append((n1, n2, wc))
        if high_corr_pairs:
            high_corr_pairs.sort(key=lambda x: abs(x[2]), reverse=True)
            print(f"\n  ! High correlation pairs (weekly >= {CORR_HIGH}):")
            for n1, n2, wc in high_corr_pairs[:10]:
                print(f"    {n1} vs {n2}: {wc:.3f}")

    # Record end time
    end_datetime = datetime.now()
    total_duration = time.time() - start_time

    print()
    print(f"{Colors.CYAN}{'=' * 60}{Colors.RESET}")
    print(f"{Colors.CYAN}Finished:{Colors.RESET}       {Colors.GREEN}{end_datetime.strftime('%Y-%m-%d %H:%M:%S')}{Colors.RESET}")
    print(f"{Colors.CYAN}Total Duration:{Colors.RESET} {Colors.GREEN}{format_duration(total_duration)}{Colors.RESET}")


if __name__ == '__main__':
    main()
